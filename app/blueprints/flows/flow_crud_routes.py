from __future__ import annotations

import json
import os
import shutil
from io import BytesIO

from flask import abort, current_app, redirect, request, send_file, url_for
from werkzeug.utils import secure_filename

from app.services.audit_service import record_audit
from app.utils import normalize_docx_output_path, parse_bool
from app.services.flow_service import parse_template_paragraphs
from app.services.user_context_service import get_actor_info

from app.services.flow_version_service import flow_versions_dir as _flow_versions_dir

from .flow_crud_blueprint import flow_crud_bp
from .flow_file_helpers import _resolve_task_file_path
from .flow_route_helpers import _touch_task_last_edit
from .flow_validation_helpers import _validate_flow_name


def _current_actor() -> dict[str, str]:
    work_id, label = get_actor_info()
    return {"work_id": work_id, "label": label}


def _record_flow_audit(action: str, task_id: str, detail: dict | None = None) -> None:
    payload = {"task_id": task_id}
    payload.update(detail or {})
    record_audit(
        action=action,
        actor=_current_actor(),
        detail=payload,
        task_id=task_id,
    )


@flow_crud_bp.post("/delete/<flow_name>", endpoint="delete_flow")
def delete_flow(task_id, flow_name):
    task_dir = os.path.join(current_app.config["TASK_FOLDER"], task_id)
    flow_dir = os.path.join(task_dir, "flows")
    path = os.path.join(flow_dir, f"{flow_name}.json")
    if os.path.exists(path):
        os.remove(path)
        versions_dir = _flow_versions_dir(flow_dir, flow_name)
        if os.path.isdir(versions_dir):
            shutil.rmtree(versions_dir, ignore_errors=True)
        _touch_task_last_edit(task_id)
        _record_flow_audit("flow_delete", task_id, {"flow": flow_name})
    return redirect(url_for("flow_builder_bp.flow_builder", task_id=task_id, fpage=request.form.get("fpage")))


@flow_crud_bp.post("/rename/<flow_name>", endpoint="rename_flow")
def rename_flow(task_id, flow_name):
    new_name = request.form.get("name", "").strip()
    name_error = _validate_flow_name(new_name)
    if name_error:
        return name_error, 400
    task_dir = os.path.join(current_app.config["TASK_FOLDER"], task_id)
    flow_dir = os.path.join(task_dir, "flows")
    old_path = os.path.join(flow_dir, f"{flow_name}.json")
    new_path = os.path.join(flow_dir, f"{new_name}.json")
    if not os.path.exists(old_path):
        abort(404)
    if os.path.exists(new_path):
        return "流程名稱已存在", 400
    os.rename(old_path, new_path)
    old_versions_dir = _flow_versions_dir(flow_dir, flow_name)
    new_versions_dir = _flow_versions_dir(flow_dir, new_name)
    if os.path.isdir(old_versions_dir):
        os.makedirs(os.path.dirname(new_versions_dir), exist_ok=True)
        if os.path.isdir(new_versions_dir):
            shutil.rmtree(new_versions_dir, ignore_errors=True)
        os.rename(old_versions_dir, new_versions_dir)
    _touch_task_last_edit(task_id)
    _record_flow_audit("flow_rename", task_id, {"flow": flow_name, "new_name": new_name})
    return redirect(url_for("flow_builder_bp.flow_builder", task_id=task_id, fpage=request.form.get("fpage")))


@flow_crud_bp.get("/export/<flow_name>", endpoint="export_flow")
def export_flow(task_id, flow_name):
    task_dir = os.path.join(current_app.config["TASK_FOLDER"], task_id)
    path = os.path.join(task_dir, "flows", f"{flow_name}.json")
    if not os.path.exists(path):
        abort(404)
    _record_flow_audit("flow_export", task_id, {"flow": flow_name, "export_type": "json"})
    return send_file(path, as_attachment=True, download_name=f"{flow_name}.json")


def _normalize_flow_source_for_mapping(raw_source: str, files_dir: str) -> str:
    text = str(raw_source or "").strip()
    if not text:
        return ""
    if os.path.isabs(text):
        abs_source = os.path.abspath(text)
        abs_files = os.path.abspath(files_dir)
        try:
            rel = os.path.relpath(abs_source, abs_files)
        except ValueError:
            rel = ""
        if rel and not rel.startswith("..") and not os.path.isabs(rel):
            return rel.replace("/", "\\")
        return os.path.basename(abs_source)
    return text.replace("/", "\\")


def _mapping_yes_no(value, default: str = "Y") -> str:
    if value is None or str(value).strip() == "":
        return default
    return "Y" if parse_bool(value, default=(default == "Y")) else "N"


def _quote_mapping_operation_text(text: str, *, protected_chars: tuple[str, ...] = ("/", "\\")) -> str:
    raw = str(text or "").strip()
    if not raw or not any(ch in raw for ch in protected_chars):
        return raw
    if '"' not in raw:
        return f'"{raw}"'
    if "'" not in raw:
        return f"'{raw}'"
    return raw


def _build_extract_target(params: dict) -> str:
    section = str(params.get("target_chapter_section") or "").strip()
    chapter_title = str(params.get("target_chapter_title") or "").strip()
    quoted_title = _quote_mapping_operation_text(chapter_title)
    if section and chapter_title:
        section_fold = section.casefold()
        title_fold = chapter_title.casefold()
        if title_fold in section_fold:
            return section
        return f"{section} {quoted_title}"
    if section:
        return section
    ref_raw = str(params.get("target_chapter_ref_raw") or "").strip()
    if ref_raw:
        return ref_raw
    return quoted_title


def _build_mapping_operation_for_visual(params: dict, *, kind: str) -> str:
    base = _build_extract_target(params)
    parts = [base] if base else []
    caption = str(params.get("target_caption_label") or "").strip()
    title_value = params.get("target_figure_title") if kind == "figure" else params.get("target_table_title")
    index_value = params.get("target_figure_index") if kind == "figure" else params.get("target_table_index")
    title = _quote_mapping_operation_text(str(title_value or "").strip(), protected_chars=("|", "/", "\\"))
    index = str(index_value or "").strip()
    if caption:
        parts.append(caption)
    if title:
        parts.append(f"title={title}")
    if index:
        parts.append(f"index={index}")
    return " | ".join([p for p in parts if p])


def _resolve_template_insert_label(params: dict, template_insert_map: dict[int, str]) -> str:
    raw_idx = str(params.get("template_index") or "").strip()
    if not raw_idx:
        return ""
    try:
        idx = int(raw_idx)
    except Exception:
        return raw_idx
    label = str(template_insert_map.get(idx) or "").strip()
    return label or str(idx)


def _mapping_template_mode_label(params: dict, template_name: str, insert_label: str) -> str:
    if not template_name:
        return ""
    raw_mode = str(params.get("template_mode") or "").strip()
    if raw_mode == "replace":
        return "取代該段落"
    if raw_mode == "insert_after" or insert_label:
        return "段落後新增"
    return ""


def _compose_template_insert_label(display: str, text: str) -> str:
    disp = str(display or "").strip()
    body = str(text or "").strip()
    if disp and body:
        disp_fold = disp.casefold()
        body_fold = body.casefold()
        if disp_fold in body_fold:
            return body
        if body_fold in disp_fold:
            return disp
        return f"{disp} {body}"
    return disp or body


def _mapping_operation_with_params(base: str, params: dict, keys: list[str]) -> str:
    parts = [base]
    for key in keys:
        if key in params:
            value = str(params.get(key) or "").strip()
            if value:
                parts.append(f"{key}={value}")
    return " | ".join(parts)


def _flow_step_to_mapping_row(
    step: dict,
    flow_output_rel: str,
    files_dir: str,
    template_name: str,
    template_insert_map: dict[int, str],
) -> dict | None:
    stype = str(step.get("type") or "").strip()
    params = dict(step.get("params") or {})
    out_path = os.path.dirname(flow_output_rel).replace("/", "\\") if flow_output_rel else ""
    out_name = os.path.basename(flow_output_rel) if flow_output_rel else ""
    insert_label = _resolve_template_insert_label(params, template_insert_map)
    insert_mode = _mapping_template_mode_label(params, template_name, insert_label)

    if stype == "extract_word_all_content":
        return {
            "source": _normalize_flow_source_for_mapping(params.get("input_file"), files_dir),
            "item_type": "All",
            "operation": "All",
            "include_title": "Y",
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype == "extract_word_chapter":
        chapter = _build_extract_target(params)
        subtitle = str(params.get("target_subtitle") or "").strip()
        operation = chapter
        if chapter and subtitle:
            operation = f"{chapter}\\{_quote_mapping_operation_text(subtitle)}"
        return {
            "source": _normalize_flow_source_for_mapping(params.get("input_file"), files_dir),
            "item_type": "",
            "operation": operation,
            "include_title": "N" if parse_bool(params.get("hide_chapter_title"), False) else "Y",
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype == "extract_specific_figure_from_word":
        return {
            "source": _normalize_flow_source_for_mapping(params.get("input_file"), files_dir),
            "item_type": "Figure Table" if parse_bool(params.get("allow_table_figure_container"), False) else "Figure",
            "operation": _build_mapping_operation_for_visual(params, kind="figure"),
            "include_title": _mapping_yes_no(params.get("include_caption"), default="Y"),
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype == "extract_specific_table_from_word":
        return {
            "source": _normalize_flow_source_for_mapping(params.get("input_file"), files_dir),
            "item_type": "Table",
            "operation": _build_mapping_operation_for_visual(params, kind="table"),
            "include_title": _mapping_yes_no(params.get("include_caption"), default="Y"),
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype == "extract_pdf_pages_as_images":
        return {
            "source": _normalize_flow_source_for_mapping(params.get("input_file"), files_dir),
            "item_type": "PDF Image",
            "operation": "",
            "include_title": "Y",
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype == "insert_image":
        return {
            "source": _normalize_flow_source_for_mapping(params.get("input_file"), files_dir),
            "item_type": "Add Image",
            "operation": _mapping_operation_with_params("Add Image", params, ["align"]),
            "include_title": "",
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype == "insert_text":
        return {
            "source": str(params.get("text") or "").strip(),
            "item_type": "Add Text",
            "operation": _mapping_operation_with_params("Add Text", params, ["align", "bold", "font_size"]),
            "include_title": "",
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype in {"insert_roman_heading", "insert_bulleted_heading", "insert_numbered_heading"}:
        heading_type_map = {
            "insert_numbered_heading": "Numbered Heading",
            "insert_roman_heading": "Roman Heading",
            "insert_bulleted_heading": "Bulleted Heading",
        }
        operation_parts = [heading_type_map[stype]]
        if stype in {"insert_numbered_heading", "insert_roman_heading"}:
            operation_parts.append(f"level={str(params.get('level', 0)).strip() or '0'}")
        if "bold" in params:
            operation_parts.append(f"bold={str(params.get('bold')).strip()}")
        if "font_size" in params:
            operation_parts.append(f"font_size={str(params.get('font_size')).strip()}")
        return {
            "source": str(params.get("text") or "").strip(),
            "item_type": heading_type_map[stype],
            "operation": " | ".join(operation_parts),
            "include_title": "",
            "out_path": out_path,
            "out_name": out_name,
            "template": template_name,
            "insert": insert_label,
            "insert_mode": insert_mode,
        }
    if stype == "copy_files":
        return {
            "source": _normalize_flow_source_for_mapping(params.get("source_dir"), files_dir),
            "item_type": "Copy File",
            "operation": str(params.get("keywords") or "").strip(),
            "include_title": "",
            "out_path": str(params.get("dest_dir") or "").strip().replace("/", "\\"),
            "out_name": str(params.get("target_name") or "").strip(),
            "template": "",
            "insert": "",
            "insert_mode": "",
        }
    if stype == "copy_directory":
        return {
            "source": _normalize_flow_source_for_mapping(params.get("source_dir"), files_dir),
            "item_type": "Copy Folder",
            "operation": str(params.get("keywords") or "").strip(),
            "include_title": "",
            "out_path": str(params.get("dest_dir") or "").strip().replace("/", "\\"),
            "out_name": str(params.get("target_name") or "").strip(),
            "template": "",
            "insert": "",
            "insert_mode": "",
        }
    return None


def _build_mapping_rows_for_flow(flow_data: dict | list, files_dir: str) -> list[dict]:
    if isinstance(flow_data, dict):
        steps = list(flow_data.get("steps") or [])
        output_filename, output_error = normalize_docx_output_path(flow_data.get("output_filename"), default="")
        if output_error:
            output_filename = ""
        template_rel = str(flow_data.get("template_file") or "").strip()
        template_name = os.path.basename(template_rel)
        template_insert_map: dict[int, str] = {}
        if template_rel:
            try:
                template_abs = _resolve_task_file_path(files_dir, template_rel, expect_dir=False)
                parsed = parse_template_paragraphs(template_abs)
                for item in parsed:
                    try:
                        idx = int(item.get("index"))
                    except Exception:
                        continue
                    display = str(item.get("display") or "").strip()
                    text = str(item.get("text") or "").strip()
                    label = _compose_template_insert_label(display, text)
                    if label and idx not in template_insert_map:
                        template_insert_map[idx] = label
            except Exception:
                template_insert_map = {}
    else:
        steps = list(flow_data or [])
        output_filename = ""
        template_name = ""
        template_insert_map = {}

    rows = []
    for step in steps:
        row = _flow_step_to_mapping_row(step, output_filename, files_dir, template_name, template_insert_map)
        if row:
            rows.append(row)
    return rows


def _build_mapping_workbook_bytes(flow_data: dict | list, files_dir: str, *, source_flow_name: str = "") -> bytes:
    return _build_mapping_workbook_from_rows(
        _build_mapping_rows_for_flow(flow_data, files_dir),
        source_flow_name=source_flow_name,
    )


def _build_mapping_workbook_from_rows(rows: list[dict], *, source_flow_name: str = "") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to export Mapping Excel") from exc

    headers = [
        "輸入檔案名稱/資料夾名稱/文字內容",
        "擷取類型",
        "擷取段落",
        "包含標題",
        "輸出資料夾名稱",
        "輸出檔案名稱",
        "模板文件",
        "插入段落名稱",
        "插入方式",
    ]
    if source_flow_name:
        headers = ["來源流程", *headers]

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping定義"
    ws.append(headers)
    # 標題欄位格式
    header_input_fill = PatternFill(fill_type="solid", fgColor="375623")
    header_output_fill = PatternFill(fill_type="solid", fgColor="C65911")
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border_style = Side(style="medium", color="000000")
    header_border = Border(left=header_border_style, right=header_border_style, top=header_border_style, bottom=header_border_style)

    data_input_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    data_output_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    data_font = Font(name="微軟正黑體", size=11, color="000000")
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_border_style = header_border_style
    data_header_border = Border(left=data_border_style, right=data_border_style, top=data_border_style, bottom=data_border_style)
    input_column_indexes = {1, 2, 3, 4, 5} if source_flow_name else {1, 2, 3, 4}

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        if col_idx in input_column_indexes:
            cell.fill = header_input_fill
        else:
            cell.fill = header_output_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
        
    for row in rows:
        row_values = [
            row["source"],
            row["item_type"],
            row["operation"],
            row["include_title"],
            row["out_path"],
            row["out_name"],
            row["template"],
            row["insert"],
            row["insert_mode"],
        ]
        if source_flow_name:
            row_values = [row.get("source_flow") or "", *row_values]
        ws.append(
            row_values
        )
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx in input_column_indexes:
                cell.fill = data_input_fill
            else:
                cell.fill = data_output_fill
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_header_border

    if source_flow_name:
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 143
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 80
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 54
        ws.column_dimensions["G"].width = 54
        ws.column_dimensions["H"].width = 54
        ws.column_dimensions["I"].width = 54
        ws.column_dimensions["J"].width = 18
    else:
        ws.column_dimensions["A"].width = 143
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 54
        ws.column_dimensions["F"].width = 54
        ws.column_dimensions["G"].width = 54
        ws.column_dimensions["H"].width = 54
        ws.column_dimensions["I"].width = 18

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@flow_crud_bp.get("/export-mapping/<flow_name>", endpoint="export_flow_mapping")
def export_flow_mapping(task_id, flow_name):
    task_dir = os.path.join(current_app.config["TASK_FOLDER"], task_id)
    flow_path = os.path.join(task_dir, "flows", f"{flow_name}.json")
    files_dir = os.path.join(task_dir, "files")
    if not os.path.exists(flow_path):
        abort(404)

    with open(flow_path, "r", encoding="utf-8") as f:
        flow_data = json.load(f)

    payload = _build_mapping_workbook_bytes(flow_data, files_dir)
    _record_flow_audit("flow_export_mapping", task_id, {"flow": flow_name, "export_type": "mapping_excel"})
    return send_file(
        BytesIO(payload),
        as_attachment=True,
        download_name=f"{flow_name}_mapping.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@flow_crud_bp.post("/export-mapping/merged", endpoint="export_merged_flow_mapping")
def export_merged_flow_mapping(task_id):
    task_dir = os.path.join(current_app.config["TASK_FOLDER"], task_id)
    flow_dir = os.path.join(task_dir, "flows")
    files_dir = os.path.join(task_dir, "files")
    selected_flow_names = [str(name or "").strip() for name in request.form.getlist("flow_names")]
    selected_flow_names = [name for name in selected_flow_names if name]
    if not selected_flow_names:
        return "請至少選擇一個流程", 400

    rows: list[dict] = []
    exported_names: list[str] = []
    for flow_name in selected_flow_names:
        flow_path = os.path.join(flow_dir, f"{flow_name}.json")
        if not os.path.isfile(flow_path):
            return f"找不到流程：{flow_name}", 404
        with open(flow_path, "r", encoding="utf-8") as f:
            flow_data = json.load(f)
        flow_rows = _build_mapping_rows_for_flow(flow_data, files_dir)
        for row in flow_rows:
            row["source_flow"] = flow_name
        rows.extend(flow_rows)
        exported_names.append(flow_name)

    if not rows:
        return "選取的流程沒有可匯出的 Mapping 步驟", 400

    payload = _build_mapping_workbook_from_rows(rows, source_flow_name="來源流程")
    _record_flow_audit(
        "flow_export_mapping_merged",
        task_id,
        {"flows": exported_names, "export_type": "mapping_excel"},
    )
    return send_file(
        BytesIO(payload),
        as_attachment=True,
        download_name="merged_mapping.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@flow_crud_bp.post("/import", endpoint="import_flow")
def import_flow(task_id):
    task_dir = os.path.join(current_app.config["TASK_FOLDER"], task_id)
    flow_dir = os.path.join(task_dir, "flows")
    os.makedirs(flow_dir, exist_ok=True)
    uploaded = request.files.get("flow_file")
    if not uploaded or not uploaded.filename.endswith(".json"):
        return "請上傳 JSON 檔", 400
    name = os.path.splitext(secure_filename(uploaded.filename))[0]
    path = os.path.join(flow_dir, f"{name}.json")
    uploaded.save(path)
    _touch_task_last_edit(task_id)
    _record_flow_audit(
        "flow_import",
        task_id,
        {
            "flow": name,
            "source_file": os.path.basename(uploaded.filename),
        },
    )
    return redirect(url_for("flow_builder_bp.flow_builder", task_id=task_id, fpage=request.form.get("fpage")))
