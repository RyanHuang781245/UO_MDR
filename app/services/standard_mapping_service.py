from __future__ import annotations

import copy
import html
import os
import re
import tempfile
import zipfile
from difflib import SequenceMatcher
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook
from modules.chapter_section_parse import parse_chapter_section_expression
from modules.docx_toc import extract_toc_entries_from_parts
from modules.extract_word_chapter import (
    build_style_heading_rank_map,
    build_style_outline_map,
    find_section_range_children,
    get_effective_heading_depth,
    is_toc_paragraph,
    iter_paragraphs,
    parse_styles_numpr,
)

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS = {"w": W_NS}
XML_NS = "http://www.w3.org/XML/1998/namespace"

EXCEL_STANDARD_COL_INDEX = 5
ISO_FAMILY_SHEETS = ["ISO", "BS-EN-DIN(歐洲國家標準)"]
RED_COLOR = "FF0000"
BLUE_COLOR = "2563EB"
STANDARD_LEVELS = ("BS EN ISO", "BS EN", "EN", "EN ISO", "BS ISO", "ISO", "BS")
EN_PRIORITY_LEVELS = ("BS EN ISO", "BS EN", "EN", "EN ISO")
DEFAULT_ISO_PRIORITY = STANDARD_LEVELS
DEFAULT_ENABLED_STANDARD_LEVELS = EN_PRIORITY_LEVELS
DEFAULT_PREFER_LATEST_EN_VARIANTS = True
AVAILABLE_HEADER_OPTIONS = (
    "Standards",
    "Issued Year",
    "EU Harmonised Standards under MDR 2017/745 (YES/NO)",
    "Title",
)
HEADER_FIELD_IDS = {
    "Standards": "standards",
    "Issued Year": "issued_year",
    "EU Harmonised Standards under MDR 2017/745 (YES/NO)": "eu_harmonised",
    "Title": "title",
}
HEADER_FIELD_IDS_REVERSE = {value: key for key, value in HEADER_FIELD_IDS.items()}
DEFAULT_REQUIRED_HEADERS = (
    "Standards",
    "Issued Year",
    "EU Harmonised Standards under MDR 2017/745 (YES/NO)",
    "Title",
)
MINIMUM_REQUIRED_HEADERS = (
    "Standards",
    "Issued Year",
)

HEADER_ALIASES = {
    "Standards": [
        "Standards",
        "Standard",
        "Standard No",
        "Standard Number",
    ],
    "Issued Year": [
        "Issued Year",
        "Issue Year",
        "Year",
    ],
    "EU Harmonised Standards under MDR 2017/745 (YES/NO)": [
        "EU Harmonised Standards",
        "EU Harmonised Standards under MDR",
        "EU Harmonised Standards under MDR 2017/745 (YES/NO)",
        "EU Harmonised Standards under MDR 2017/745(YES/NO)",
        "EU Harmonized Standards under MDR 2017/745 (YES/NO)",
        "EU Harmonized Standards under MDR 2017/745(YES/NO)",
        "EU Harmonised Standards under MDR 2017/745",
        "EU Harmonized Standards under MDR 2017/745",
    ],
    "Title": [
        "Title",
        "Standard Title",
    ],
}

HEADER_KEYWORDS = {
    "Standards": {"STANDARD"},
    "Issued Year": {"ISSUED", "YEAR"},
    "EU Harmonised Standards under MDR 2017/745 (YES/NO)": {"EU", "HARMONISED", "MDR", "2017/745"},
    "Title": {"TITLE"},
}

EXCEL_TITLE_HEADER_ALIASES = ("標準名稱", "TITLE", "STANDARD TITLE")
HARMONISED_REFERENCE_HEADERS = ("Reference and title Provision",)


def qn(tag: str) -> str:
    prefix, local = tag.split(":")
    if prefix != "w":
        raise ValueError("qn() only supports w: namespace")
    return f"{{{W_NS}}}{local}"


def normalize_text(text: str) -> str:
    if text is None:
        return ""
    text = str(text).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_standard_text(text: str) -> str:
    return normalize_text(text).replace("：", ":").replace("／", "/").replace("＋", "+")


def normalize_harmonised_identifier(text: str) -> str:
    normalized = normalize_standard_text(text).upper()
    collapsed = re.sub(r"\s*([:/+()-])\s*", r"\1", normalized)
    collapsed = re.sub(r"\s+", " ", collapsed).strip()
    return collapsed.strip()


def normalize_harmonised_standard_text(text: str, title: str = "") -> str:
    raw_text = "" if text is None else str(text).replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in raw_text.split("\n") if line.strip()]
    first_line = lines[0] if lines else ""
    return normalize_harmonised_identifier(first_line)


def extract_harmonised_reference_entries(text: str) -> list[str]:
    raw_text = "" if text is None else str(text).replace("\r\n", "\n").replace("\r", "\n")
    entries: list[str] = []
    for line in raw_text.split("\n"):
        normalized_line = normalize_harmonised_identifier(line)
        if not normalized_line:
            continue
        if re.match(r"^(?:EN|ISO|BS|DIN|IEC|ASTM)\b", normalized_line):
            entries.append(normalized_line)
    return entries


def load_harmonised_reference_index(reference_path: str | os.PathLike | None = None) -> set[str]:
    path = Path(reference_path or "download.xlsx")
    if not path.is_file():
        return set()
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        lookup: set[str] = set()
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True))
            col_index = find_excel_header_col_index(rows, HARMONISED_REFERENCE_HEADERS)
            if col_index is None:
                continue
            for row in ws.iter_rows(min_row=1, values_only=True):
                values = list(row or [])
                value = values[col_index] if col_index < len(values) else None
                for normalized in extract_harmonised_reference_entries(value):
                    if normalized != normalize_harmonised_standard_text(HARMONISED_REFERENCE_HEADERS[0]):
                        lookup.add(normalized)
        return lookup
    finally:
        wb.close()


def normalize_regulation_lookup_key(text: str) -> str:
    normalized = normalize_text(text).upper()
    if not normalized:
        return ""
    normalized = normalized.replace("–", "-").replace("—", "-").replace("：", ":").replace("／", "/")
    normalized = re.sub(r"\bREVISION\b", "REV", normalized)
    normalized = re.sub(r"\bREV\.\b", "REV", normalized)
    normalized = re.sub(r"\bREV\.\s*", "REV ", normalized)
    return re.sub(r"[^A-Z0-9]+", "", normalized)


def is_regulation_lookup_target(text: str) -> bool:
    normalized_key = normalize_regulation_lookup_key(text)
    return normalized_key.startswith(("MEDDEV", "MDCG"))


def extract_year_from_regulation_date(value) -> int | None:
    if value is None:
        return None
    if hasattr(value, "year") and getattr(value, "year", None):
        try:
            return int(value.year)
        except Exception:
            return None
    years = re.findall(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", normalize_text(value))
    if years:
        return max(int(year) for year in years)
    return None


def load_regulation_reference_index(reference_path: str | os.PathLike | None = None) -> dict[str, list[dict]]:
    path = Path(reference_path or "")
    if not path.is_file():
        return {}
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        index: dict[str, list[dict]] = {}
        for ws in wb.worksheets:
            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                values = list(row or [])
                key_text = normalize_text(values[4] if len(values) > 4 else "")
                if not key_text:
                    continue
                normalized_key = normalize_regulation_lookup_key(key_text)
                if not normalized_key:
                    continue
                date_value = values[6] if len(values) > 6 else None
                latest_year = extract_year_from_regulation_date(date_value)
                title_text = normalize_text(values[5] if len(values) > 5 else "")
                index.setdefault(normalized_key, []).append({
                    "sheet_name": ws.title,
                    "excel_col_letter": "E",
                    "excel_row_index": row_idx,
                    "matched_standard_no": key_text,
                    "matched_display_standard_no": key_text,
                    "matched_title": title_text,
                    "candidate_harmonised": "",
                    "latest_year": latest_year,
                    "standard_level": "REGULATION",
                    "standard_level_rank": 0,
                    "search_family": "REGULATION",
                    "apply_year_comparison": True,
                    "decision": "kept",
                    "decision_reason": "命中國家法規條文登記表候選",
                    "candidate_id": "",
                })
        return index
    finally:
        wb.close()


def is_harmonised_standard(std_no: str, title: str, harmonised_reference_index: set[str] | None) -> bool:
    if not harmonised_reference_index:
        return False
    normalized = normalize_harmonised_standard_text(std_no)
    return bool(normalized and normalized in harmonised_reference_index)


def normalize_key_for_search(text: str) -> str:
    text = normalize_text(text).upper()
    text = text.replace("–", "-").replace("—", "-")
    text = text.replace("：", ":").replace("／", "/")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*([\(\)])\s*", r"\1", text)
    return text


def compact_key_for_search(text: str) -> str:
    normalized = normalize_key_for_search(text)
    return re.sub(r"[^A-Z0-9]+", "", normalized)


def text_matches_manual_header(cell_text: str, target_text: str) -> bool:
    normalized_cell = normalize_key_for_search(cell_text)
    normalized_target = normalize_key_for_search(target_text)
    if normalized_cell and normalized_target and normalized_cell == normalized_target:
        return True

    compact_cell = compact_key_for_search(cell_text)
    compact_target = compact_key_for_search(target_text)
    if compact_cell and compact_target and compact_cell == compact_target:
        return True

    return False


def find_excel_title_col_index(rows: list[tuple | list]) -> int | None:
    for row in rows[:10]:
        values = list(row or [])
        for index, value in enumerate(values):
            cell_text = normalize_text(value)
            if not cell_text:
                continue
            if any(text_matches_manual_header(cell_text, alias) for alias in EXCEL_TITLE_HEADER_ALIASES):
                return index
    return None


def find_excel_header_col_index(rows: list[tuple | list], aliases: tuple[str, ...]) -> int | None:
    for row in rows[:10]:
        values = list(row or [])
        for index, value in enumerate(values):
            cell_text = normalize_text(value)
            if not cell_text:
                continue
            if any(text_matches_manual_header(cell_text, alias) for alias in aliases):
                return index
    return None


def header_matches_target(header_text: str, target_name: str) -> bool:
    normalized_header = normalize_key_for_search(header_text)
    compact_header = compact_key_for_search(header_text)
    aliases = HEADER_ALIASES.get(target_name, [target_name])

    for alias in aliases:
        if normalized_header == normalize_key_for_search(alias):
            return True
        if compact_header == compact_key_for_search(alias):
            return True

    keywords = HEADER_KEYWORDS.get(target_name)
    if keywords:
        header_words = set(re.findall(r"[A-Z0-9/]+", normalized_header))
        if keywords.issubset(header_words):
            return True

    return False


def detect_search_family(standard_name: str) -> str | None:
    s = normalize_key_for_search(standard_name)
    if not s:
        return None
    if "ASTM" in s:
        return "ASTM"
    if re.match(r"^(?:IEC)\b", s):
        return "IEC_FAMILY"
    if re.match(r"^(?:BS\s*EN\s*ISO|DIN\s*EN\s*ISO|EN\s*ISO|BS\s*ISO|ISO|BS\s*EN|DIN\s*EN|EN|BS)(?=\s|\d|\b)", s):
        return "ISO_FAMILY"
    return None


def classify_standard_level(std_no: str) -> tuple[str, int]:
    s = normalize_key_for_search(std_no)
    if re.match(r"^BS\s*EN\s*ISO(?=\s|\d|\b)", s):
        return "BS EN ISO", 7
    if re.match(r"^(?:BS\s*EN|DIN\s*EN)(?=\s|\d|\b)", s):
        return "BS EN", 6
    if re.match(r"^EN(?=\s|\d|\b)", s):
        return "EN", 5
    if re.match(r"^(?:DIN\s*EN\s*ISO|EN\s*ISO)(?=\s|\d|\b)", s):
        return "EN ISO", 4
    if re.match(r"^BS\s*ISO(?=\s|\d|\b)", s):
        return "BS ISO", 3
    if re.match(r"^ISO(?=\s|\d|\b)", s):
        return "ISO", 2
    if re.match(r"^BS(?=\s|\d|\b)", s):
        return "BS", 1
    if re.match(r"^IEC(?=\s|\d|\b)", s):
        return "IEC", 1
    return "OTHER", 0


def normalize_iso_priority(priority_order: list[str] | tuple[str, ...] | None) -> tuple[str, ...]:
    if not priority_order:
        return DEFAULT_ISO_PRIORITY
    normalized = []
    seen = set()
    for item in priority_order:
        value = normalize_key_for_search(item)
        if value == "BS EN ISO":
            label = "BS EN ISO"
        elif value == "BS EN":
            label = "BS EN"
        elif value == "EN":
            label = "EN"
        elif value == "EN ISO":
            label = "EN ISO"
        elif value == "BS ISO":
            label = "BS ISO"
        elif value == "ISO":
            label = "ISO"
        elif value == "BS":
            label = "BS"
        else:
            continue
        if label not in seen:
            normalized.append(label)
            seen.add(label)
    for label in DEFAULT_ISO_PRIORITY:
        if label not in seen:
            normalized.append(label)
    return tuple(normalized[: len(DEFAULT_ISO_PRIORITY)])


def normalize_enabled_standard_levels(levels: list[str] | tuple[str, ...] | None) -> tuple[str, ...]:
    if levels is None:
        return DEFAULT_ENABLED_STANDARD_LEVELS
    normalized = []
    seen = set()
    for item in levels:
        value = normalize_key_for_search(item)
        if value == "BS EN ISO":
            label = "BS EN ISO"
        elif value == "BS EN":
            label = "BS EN"
        elif value == "EN":
            label = "EN"
        elif value == "EN ISO":
            label = "EN ISO"
        elif value == "BS ISO":
            label = "BS ISO"
        elif value == "ISO":
            label = "ISO"
        elif value == "BS":
            label = "BS"
        else:
            continue
        if label not in seen:
            normalized.append(label)
            seen.add(label)
    return tuple(normalized)


def normalize_required_headers(required_headers: list[str] | tuple[str, ...] | None) -> tuple[str, ...]:
    if not required_headers:
        return DEFAULT_REQUIRED_HEADERS
    normalized = []
    seen = set()
    for item in required_headers:
        for header_name in AVAILABLE_HEADER_OPTIONS:
            if header_matches_target(item, header_name) and header_name not in seen:
                normalized.append(header_name)
                seen.add(header_name)
                break
    if not normalized:
        return DEFAULT_REQUIRED_HEADERS
    return tuple(normalized)


def normalize_manual_header_mappings(
    manual_header_mappings: dict[int | str, dict[str, str]] | None,
) -> dict[int, dict[str, str]]:
    normalized: dict[int, dict[str, str]] = {}
    if not manual_header_mappings:
        return normalized

    for raw_table_index, raw_mapping in manual_header_mappings.items():
        try:
            table_index = int(raw_table_index)
        except (TypeError, ValueError):
            continue
        if not isinstance(raw_mapping, dict):
            continue

        mapping: dict[str, str] = {}
        for raw_key, raw_value in raw_mapping.items():
            key_text = normalize_text(raw_key)
            value_text = normalize_text(raw_value)
            if not value_text:
                continue

            target_name = HEADER_FIELD_IDS_REVERSE.get(key_text)
            if target_name is None:
                for header_name in AVAILABLE_HEADER_OPTIONS:
                    if header_matches_target(key_text, header_name):
                        target_name = header_name
                        break
            if target_name is None:
                continue
            mapping[target_name] = value_text

        if mapping:
            normalized[table_index] = mapping

    return normalized


def extract_iso_family_core(std_no: str) -> str:
    s = normalize_key_for_search(std_no)
    if not s:
        return ""
    s = re.sub(r"/A\d+(?::\s*(19\d{2}|20\d{2}))?.*$", "", s).strip()
    s = re.sub(r":\s*(19\d{2}|20\d{2}).*$", "", s).strip()
    return re.sub(r"^(?:BS\s*EN\s*ISO|DIN\s*EN\s*ISO|EN\s*ISO|BS\s*EN|DIN\s*EN|EN|BS\s*ISO|ISO|BS|IEC)\s*", "", s).strip()


def extract_display_standard_no(std_no: str) -> str:
    family = detect_search_family(std_no)
    s = normalize_text(std_no)
    if family in {"ISO_FAMILY", "IEC_FAMILY"}:
        s = s.replace("：", ":")
        return re.sub(r"\s*:\s*(19\d{2}|20\d{2}).*$", "", s).strip()
    if family == "ASTM":
        return re.sub(r"\s*-\s*(\d{2}[A-Z]?)(?!\d).*$", "", s).strip()
    return s


def make_row_key(table_index: int, row_index: int) -> str:
    return f"table-{table_index}-row-{row_index}"


def make_candidate_id(candidate: dict) -> str:
    return "|".join([
        normalize_text(candidate.get("sheet_name", "")),
        str(candidate.get("excel_row_index", "")),
        normalize_text(candidate.get("matched_standard_no", "")),
    ])


def get_all_text(node: etree._Element) -> str:
    texts = node.xpath(".//w:t/text()", namespaces=NS)
    return normalize_text("".join(texts))


def get_grid_span(tc: etree._Element) -> int:
    vals = tc.xpath("./w:tcPr/w:gridSpan/@w:val", namespaces=NS)
    if vals:
        try:
            return int(vals[0])
        except Exception:
            return 1
    return 1


def ensure_cell_has_text_node(tc: etree._Element) -> etree._Element:
    t_nodes = tc.xpath(".//w:t", namespaces=NS)
    if t_nodes:
        return t_nodes[0]
    p = tc.find("w:p", namespaces=NS)
    if p is None:
        p = etree.SubElement(tc, qn("w:p"))
    r = p.find("w:r", namespaces=NS)
    if r is None:
        r = etree.SubElement(p, qn("w:r"))
    return etree.SubElement(r, qn("w:t"))


def merge_text_segments(segments: list[tuple[str, bool]]) -> list[tuple[str, bool]]:
    merged: list[tuple[str, bool]] = []
    for text, is_red in segments:
        if not text:
            continue
        if merged and merged[-1][1] == is_red:
            merged[-1] = (merged[-1][0] + text, is_red)
        else:
            merged.append((text, is_red))
    return merged or [("", False)]


def build_diff_segments(old_text: str, new_text: str) -> list[tuple[str, bool]]:
    old_text = "" if old_text is None else str(old_text)
    new_text = "" if new_text is None else str(new_text)
    if old_text == new_text:
        return [(new_text, False)]
    if old_text and new_text.endswith(old_text):
        prefix = new_text[:-len(old_text)]
        prefix_core = prefix.rstrip()
        prefix_space = prefix[len(prefix_core):]
        return merge_text_segments([
            (prefix_core, True),
            (prefix_space, False),
            (old_text, False),
        ])
    matcher = SequenceMatcher(a=old_text, b=new_text)
    segments = []
    for tag, _, _, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            segments.append((new_text[j1:j2], False))
        elif tag in {"replace", "insert"}:
            segments.append((new_text[j1:j2], True))
    return merge_text_segments(segments)


def build_year_segments(old_text: str, new_text: str) -> list[tuple[str, bool]]:
    old_text = normalize_text(old_text)
    new_text = normalize_text(new_text)
    if old_text == new_text:
        return [(new_text, False)]
    if (
        re.fullmatch(r"(19|20)\d{2}", old_text)
        and re.fullmatch(r"(19|20)\d{2}", new_text)
        and old_text[:2] == new_text[:2]
    ):
        return merge_text_segments([(new_text[:2], False), (new_text[2:], True)])
    return build_diff_segments(old_text, new_text)


def build_preserve_original_segments(old_text: str, new_text: str) -> list[tuple[str, bool]]:
    old_text = normalize_text(old_text)
    new_text = normalize_text(new_text)
    if old_text == new_text:
        return [(new_text, False)]
    if not old_text:
        return [(new_text, True)]
    if not new_text:
        return [(old_text, False)]
    return merge_text_segments([
        (old_text, False),
        (" ", False),
        (new_text, True),
    ])


def get_first_run_properties(tc: etree._Element) -> etree._Element | None:
    rpr = tc.find(".//w:r/w:rPr", namespaces=NS)
    return copy.deepcopy(rpr) if rpr is not None else None


def set_run_color(run: etree._Element, color_hex: str):
    rpr = run.find("w:rPr", namespaces=NS)
    if rpr is None:
        rpr = etree.SubElement(run, qn("w:rPr"))
    color = rpr.find("w:color", namespaces=NS)
    if color is None:
        color = etree.SubElement(rpr, qn("w:color"))
    color.set(qn("w:val"), color_hex)


def rebuild_cell_with_segments(tc: etree._Element, segments: list[tuple[str, bool]]):
    segments = merge_text_segments(segments)
    p_nodes = tc.findall("w:p", namespaces=NS)
    paragraph = p_nodes[0] if p_nodes else etree.SubElement(tc, qn("w:p"))
    template_rpr = get_first_run_properties(tc)

    for child in list(tc):
        if child.tag == qn("w:tcPr") or child is paragraph:
            continue
        tc.remove(child)

    for child in list(paragraph):
        if child.tag != qn("w:pPr"):
            paragraph.remove(child)

    for text, is_red in segments:
        run = etree.SubElement(paragraph, qn("w:r"))
        if template_rpr is not None:
            run.append(copy.deepcopy(template_rpr))
        if is_red:
            set_run_color(run, RED_COLOR)
        t = etree.SubElement(run, qn("w:t"))
        if text != text.strip() or "  " in text:
            t.set(f"{{{XML_NS}}}space", "preserve")
        t.text = text


def rebuild_cell_with_single_color(tc: etree._Element, text: str, color_hex: str):
    rebuild_cell_with_segments(tc, [(normalize_text(text), False)])
    paragraph = tc.find("w:p", namespaces=NS)
    if paragraph is None:
        return
    first_run = paragraph.find("w:r", namespaces=NS)
    if first_run is not None:
        set_run_color(first_run, color_hex)


def get_run_color(run: etree._Element) -> str:
    vals = run.xpath("./w:rPr/w:color/@w:val", namespaces=NS)
    if not vals:
        return ""
    color = normalize_text(vals[0]).upper()
    return "" if color == "AUTO" else color


def extract_cell_runs(tc: etree._Element) -> list[tuple[str, str]]:
    segments: list[tuple[str, str]] = []
    paragraphs = tc.xpath("./w:p", namespaces=NS)
    for p_idx, paragraph in enumerate(paragraphs):
        if p_idx > 0:
            segments.append(("\n", ""))
        runs = paragraph.xpath("./w:r", namespaces=NS)
        for run in runs:
            text = "".join(run.xpath(".//w:t/text()", namespaces=NS))
            if text:
                segments.append((text, get_run_color(run)))
    if segments:
        return segments
    plain_text = get_all_text(tc)
    return [(plain_text, "")] if plain_text else []


def format_cell_runs_as_html(tc: etree._Element) -> str:
    parts = []
    for text, color in extract_cell_runs(tc):
        escaped = html.escape(text).replace("\n", "<br>")
        if color:
            parts.append(f'<span style="color: #{color.lower()};">{escaped}</span>')
        else:
            parts.append(escaped)
    return "".join(parts) or "&nbsp;"


def unzip_docx(docx_path: str, extract_dir: str):
    with zipfile.ZipFile(docx_path, "r") as archive:
        archive.extractall(extract_dir)


def zip_to_docx(folder_path: str, output_docx_path: str):
    with zipfile.ZipFile(output_docx_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for root, _, files in os.walk(folder_path):
            for file in files:
                abs_path = os.path.join(root, file)
                rel_path = os.path.relpath(abs_path, folder_path)
                archive.write(abs_path, rel_path)


def detect_sheet_type(standard_name: str) -> str | None:
    s = normalize_key_for_search(standard_name)
    if "ASTM" in s:
        return "ASTM"
    if "IEC" in s:
        return "ISO"
    if "EN ISO" in s or "BS EN" in s or re.search(r"\bEN\b", s) or "BS ISO" in s or re.search(r"^BS\b", s):
        return "BS-EN-DIN(歐洲國家標準)"
    if re.search(r"\bISO\b", s):
        return "ISO"
    return None


def extract_latest_year_from_en_iso_style(std_no: str) -> int | None:
    years = re.findall(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", normalize_text(std_no))
    if not years:
        return None
    return max(int(y) for y in years)


def astm_two_digit_to_full_year(two_digit: int) -> int:
    return 2000 + two_digit if 0 <= two_digit <= 49 else 1900 + two_digit


def normalize_astm_standard_text(std_no: str) -> str:
    text = normalize_text(std_no).upper()
    text = text.replace("–", "-").replace("—", "-")
    return re.sub(r"\s*-\s*", "-", text)


def extract_latest_year_from_astm_style(std_no: str) -> int | None:
    normalized = normalize_astm_standard_text(std_no)
    bracket_years = re.findall(r"\((19\d{2}|20\d{2})\)", normalized)
    if bracket_years:
        return max(int(year) for year in bracket_years)
    matches = re.findall(r"-(\d{2}[A-Z]?)(?!\d)", normalized)
    if not matches:
        return None
    years = [astm_two_digit_to_full_year(int(re.match(r"\d{2}", x).group(0))) for x in matches]
    return max(years) if years else None


def extract_amendment_number(std_no: str) -> str:
    normalized = normalize_standard_text(std_no)
    match = re.search(r"(?:\+|/)\s*A(?:MD)?\s*(\d+)\b", normalized, flags=re.IGNORECASE)
    return match.group(1) if match else ""


def build_title_with_amendment(title: str, std_no: str) -> str:
    base_title = normalize_text(title)
    amendment_number = extract_amendment_number(std_no)
    if not amendment_number:
        return base_title
    suffix = f" - Amendment {amendment_number}"
    if base_title.endswith(suffix):
        return base_title
    return f"{base_title}{suffix}" if base_title else f"Amendment {amendment_number}"


def extract_standard_match_key(std_no: str, sheet_name: str) -> str:
    family = detect_search_family(std_no)
    s = normalize_key_for_search(std_no)
    if not s:
        return ""
    if family in {"ISO_FAMILY", "IEC_FAMILY"}:
        return extract_iso_family_core(std_no)
    if family == "ASTM":
        return re.sub(r"\s*-\s*(\d{2}[A-Z]?)(?!\d).*$", "", s).strip()
    s = re.sub(r":\s*(19\d{2}|20\d{2}).*$", "", s).strip()
    if sheet_name == "BS-EN-DIN(歐洲國家標準)":
        s = re.sub(r"^(?:BS|DIN)\s+(?=EN\b)", "", s).strip()
    return s


def build_sheet_records(ws, std_col_index: int = EXCEL_STANDARD_COL_INDEX) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    title_col_index = find_excel_header_col_index(rows, EXCEL_TITLE_HEADER_ALIASES)
    records = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        values = list(row)
        std_val = values[std_col_index] if std_col_index < len(values) else None
        title_val = values[title_col_index] if title_col_index is not None and title_col_index < len(values) else None
        if std_val is None:
            continue
        std_val = normalize_text(std_val)
        if not std_val:
            continue
        title_val = normalize_text(title_val)
        standard_level, standard_level_rank = classify_standard_level(std_val)
        records.append({
            "sheet_name": ws.title,
            "excel_row_index": row_idx,
            "excel_col_letter": "F",
            "standard_no": std_val,
            "standard_title": title_val,
            "standard_match_key": extract_standard_match_key(std_val, ws.title),
            "search_family": detect_search_family(std_val),
            "standard_display_no": extract_display_standard_no(std_val),
            "standard_level": standard_level,
            "standard_level_rank": standard_level_rank,
        })
    return records


def load_excel_index(excel_path: str) -> dict:
    wb = load_workbook(excel_path, data_only=True)
    needed_sheets = ["BS-EN-DIN(歐洲國家標準)", "ISO", "ASTM"]
    index = {}
    for sheet_name in needed_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Excel 缺少工作表: {sheet_name}")
        index[sheet_name] = build_sheet_records(wb[sheet_name], EXCEL_STANDARD_COL_INDEX)
    return index


def find_latest_year_from_excel(
    standard_name: str,
    excel_index: dict,
    iso_priority: list[str] | tuple[str, ...] | None = None,
    enabled_standard_levels: list[str] | tuple[str, ...] | None = None,
    harmonised_reference_index: set[str] | None = None,
    prefer_latest_en_variants: bool = DEFAULT_PREFER_LATEST_EN_VARIANTS,
) -> dict | None:
    family = detect_search_family(standard_name)
    if not family:
        return None
    query_key = extract_standard_match_key(standard_name, "")
    if not query_key:
        return None

    candidates = []
    normalized_iso_priority = normalize_iso_priority(iso_priority)
    normalized_enabled_levels = normalize_enabled_standard_levels(enabled_standard_levels)
    if family in {"ISO_FAMILY", "IEC_FAMILY"}:
        target_sheets = ISO_FAMILY_SHEETS
    elif family == "ASTM":
        target_sheets = ["ASTM"]
    else:
        sheet_name = detect_sheet_type(standard_name)
        if not sheet_name:
            return None
        target_sheets = [sheet_name]

    for sheet_name in target_sheets:
        for rec in excel_index.get(sheet_name, []):
            if query_key != rec["standard_match_key"]:
                continue
            year = (
                extract_latest_year_from_astm_style(rec["standard_no"])
                if family == "ASTM"
                else extract_latest_year_from_en_iso_style(rec["standard_no"])
            )
            candidates.append({
                "sheet_name": sheet_name,
                "excel_col_letter": "F",
                "excel_row_index": rec["excel_row_index"],
                "matched_standard_no": rec["standard_no"],
                "matched_display_standard_no": rec["standard_display_no"],
                "matched_title": build_title_with_amendment(rec.get("standard_title", ""), rec["standard_no"]),
                "candidate_harmonised": "Yes" if is_harmonised_standard(rec["standard_no"], rec.get("standard_title", ""), harmonised_reference_index) else "No",
                "latest_year": year,
                "standard_level": rec["standard_level"],
                "standard_level_rank": rec["standard_level_rank"],
                "search_family": rec["search_family"],
                "apply_year_comparison": (
                    rec["standard_level"] in normalized_enabled_levels
                    if family in {"ISO_FAMILY", "IEC_FAMILY"}
                    else True
                ),
                "decision": "kept",
                "decision_reason": "納入初始候選",
                "candidate_id": "",
            })

    if not candidates:
        return None

    all_candidates = [dict(item) for item in candidates]

    def same_type_sort_key(candidate: dict) -> tuple:
        return (candidate.get("latest_year") or 0, -len(candidate.get("matched_standard_no", "")))

    if family in {"ISO_FAMILY", "IEC_FAMILY"}:
        priority_index = {label: idx for idx, label in enumerate(normalized_iso_priority)}
        best_by_level: dict[str, dict] = {}
        for candidate in all_candidates:
          level = candidate["standard_level"]
          current = best_by_level.get(level)
          if current is None or same_type_sort_key(candidate) > same_type_sort_key(current):
              best_by_level[level] = candidate

        checked_candidates = [
            candidate
            for candidate in best_by_level.values()
            if candidate["standard_level"] in normalized_iso_priority and candidate.get("apply_year_comparison")
        ]
        fallback_candidates = [
            candidate
            for candidate in best_by_level.values()
            if candidate["standard_level"] in normalized_iso_priority and not candidate.get("apply_year_comparison")
        ]
        prioritized_best_candidates = [
            candidate for candidate in best_by_level.values() if candidate["standard_level"] in normalized_iso_priority
        ]

        if checked_candidates:
            candidates = checked_candidates
        elif prioritized_best_candidates:
            candidates = prioritized_best_candidates
        else:
            candidates = list(best_by_level.values())
        has_checked_candidates = bool(checked_candidates)

        for candidate in all_candidates:
            level = candidate["standard_level"]
            if level not in normalized_iso_priority:
                candidate["decision_reason"] = "不在優先級清單內，僅在無更高優先候選時作為後備"
            elif best_by_level.get(level) is not candidate:
                candidate["decision_reason"] = "同類型已有較新年份候選，未納入最終決選"
            elif has_checked_candidates and candidate.get("apply_year_comparison"):
                candidate["decision_reason"] = "此類型已勾選，會納入年份比較，再依優先級決選"
            elif has_checked_candidates:
                candidate["decision_reason"] = "已有勾選類型候選，此類型不參與年份比較，僅作後備"
            else:
                candidate["decision_reason"] = "目前沒有勾選候選命中，改依優先級決選"

        def candidate_sort_key(candidate: dict) -> tuple:
            latest_year = candidate.get("latest_year") or 0
            name_length = len(candidate.get("matched_standard_no", ""))
            priority_rank = len(normalized_iso_priority) - priority_index.get(candidate["standard_level"], len(normalized_iso_priority))
            prioritized = 1 if candidate["standard_level"] in normalized_iso_priority else 0
            if has_checked_candidates:
                return (latest_year, priority_rank, prioritized, -name_length)
            return (priority_rank, prioritized, -name_length, latest_year)
    else:
        for candidate in all_candidates:
            candidate["decision_reason"] = "符合查詢條件，進入最終排序"

        def candidate_sort_key(candidate: dict) -> tuple:
            latest_year = candidate.get("latest_year") or 0
            name_length = len(candidate.get("matched_standard_no", ""))
            return (1, 0, latest_year, name_length)

    if not candidates:
        return None

    candidates.sort(key=candidate_sort_key, reverse=True)
    selected = candidates[0]
    for candidate in candidates[1:]:
        if candidate.get("decision") != "excluded":
            candidate["decision"] = "kept"
            candidate["decision_reason"] = "通過篩選，但排序結果未被選用"
    selected["decision"] = "selected"
    if family in {"ISO_FAMILY", "IEC_FAMILY"}:
        if selected.get("apply_year_comparison"):
            selected["decision_reason"] = "最終採用：此類型已勾選，依年份比較後再按優先級選中"
        else:
            selected["decision_reason"] = "最終採用：此類型未勾選，因此僅依優先級選中"
    else:
        selected["decision_reason"] = "最終採用：依優先級與年份排序後選中"

    for candidate in all_candidates:
        candidate["candidate_id"] = make_candidate_id(candidate)
    ordered_candidates = sorted(
        all_candidates,
        key=lambda x: (
            1 if x.get("decision") == "selected" else 0,
            1 if x.get("decision") == "kept" else 0,
            *candidate_sort_key(x),
        ),
        reverse=True,
    )
    result = dict(selected)
    result["all_candidates"] = ordered_candidates
    result["selected_candidate_id"] = make_candidate_id(selected)
    result["auto_selected_candidate_id"] = make_candidate_id(selected)
    result["matched_harmonised"] = "Yes" if is_harmonised_standard(
        selected["matched_standard_no"],
        selected.get("matched_title", ""),
        harmonised_reference_index,
    ) else "No"
    result["iso_priority"] = list(normalized_iso_priority)
    result["enabled_standard_levels"] = list(normalized_enabled_levels)
    result["prefer_latest_en_variants"] = prefer_latest_en_variants
    return result


def find_latest_year_from_regulation_reference(
    standard_name: str,
    regulation_index: dict[str, list[dict]] | None,
) -> dict | None:
    if not regulation_index:
        return None
    query_key = normalize_regulation_lookup_key(standard_name)
    if not query_key:
        return None
    candidates = [dict(item) for item in regulation_index.get(query_key, [])]
    if not candidates:
        return None

    def candidate_sort_key(candidate: dict) -> tuple:
        return (candidate.get("latest_year") or 0, -len(candidate.get("matched_standard_no", "")))

    candidates.sort(key=candidate_sort_key, reverse=True)
    selected = candidates[0]
    for candidate in candidates[1:]:
        candidate["decision"] = "kept"
        candidate["decision_reason"] = "條文號相符，但公告年份較舊，未被選用"
    selected["decision"] = "selected"
    selected["decision_reason"] = "最終採用：命中國家法規條文登記表，依公告年份排序後選中"

    for candidate in candidates:
        candidate["candidate_id"] = make_candidate_id(candidate)

    result = dict(selected)
    result["all_candidates"] = candidates
    result["selected_candidate_id"] = selected["candidate_id"]
    result["auto_selected_candidate_id"] = selected["candidate_id"]
    result["matched_harmonised"] = ""
    result["iso_priority"] = []
    result["enabled_standard_levels"] = []
    result["prefer_latest_en_variants"] = False
    return result


def apply_candidate_override(match_info: dict, override_candidate_id: str | None) -> dict:
    result = copy.deepcopy(match_info)
    candidates = result.get("all_candidates") or []
    if not candidates:
        return result
    auto_selected_id = result.get("auto_selected_candidate_id") or make_candidate_id(result)
    selected = None
    for candidate in candidates:
        candidate["candidate_id"] = candidate.get("candidate_id") or make_candidate_id(candidate)
        if candidate["candidate_id"] == override_candidate_id:
            selected = candidate
    if selected is None:
        result["selected_candidate_id"] = auto_selected_id
        result["auto_selected_candidate_id"] = auto_selected_id
        return result
    for candidate in candidates:
        if candidate["candidate_id"] == selected["candidate_id"]:
            candidate["decision"] = "selected"
            candidate["decision_reason"] = (
                "最終採用：依自動規則選中"
                if candidate["candidate_id"] == auto_selected_id
                else "人工覆寫：使用者改選此候選"
            )
        elif candidate["candidate_id"] == auto_selected_id and selected["candidate_id"] != auto_selected_id:
            candidate["decision"] = "kept"
            candidate["decision_reason"] = "自動規則原本選用，但已被人工覆寫"
        elif candidate.get("decision") != "excluded":
            candidate["decision"] = "kept"
            candidate["decision_reason"] = "通過篩選，但未被人工選用"
    for key, value in selected.items():
        if key != "all_candidates":
            result[key] = value
    result["matched_harmonised"] = selected.get("candidate_harmonised", result.get("matched_harmonised", ""))
    result["matched_title"] = selected.get("matched_title", result.get("matched_title", ""))
    result["latest_year"] = selected.get("latest_year", result.get("latest_year"))
    result["apply_year_comparison"] = selected.get("latest_year") not in {None, ""}
    result["all_candidates"] = candidates
    result["selected_candidate_id"] = selected["candidate_id"]
    result["auto_selected_candidate_id"] = auto_selected_id
    result["manually_overridden"] = selected["candidate_id"] != auto_selected_id
    return result


def parse_table_rows(tbl: etree._Element) -> list[list[dict]]:
    parsed_rows = []
    for tr_idx, tr in enumerate(tbl.xpath("./w:tr", namespaces=NS)):
        row_items = []
        logical_col = 0
        for tc_idx, tc in enumerate(tr.xpath("./w:tc", namespaces=NS)):
            span = get_grid_span(tc)
            row_items.append({
                "tr": tr,
                "tr_idx": tr_idx,
                "tc": tc,
                "tc_idx": tc_idx,
                "text": get_all_text(tc),
                "logical_col_start": logical_col,
                "logical_col_end": logical_col + span - 1,
                "grid_span": span,
            })
            logical_col += span
        parsed_rows.append(row_items)
    return parsed_rows


def _clean_heading_title(text: str) -> str:
    return normalize_text(re.sub(r"^[\.\-:：\s]+", "", text or ""))


def _build_section_option(number: str, title: str, *, level: int | None = None, source: str = "") -> dict | None:
    clean_number = normalize_text(number).rstrip(".")
    clean_title = _clean_heading_title(title)
    value = normalize_text(f"{clean_number} {clean_title}".strip() or clean_title or clean_number)
    if not value:
        return None
    return {
        "value": value,
        "label": value,
        "number": clean_number,
        "title": clean_title,
        "level": level,
        "source": source,
    }


def inspect_document_sections(word_path: str) -> list[dict]:
    with zipfile.ZipFile(word_path, "r") as zin:
        file_map = {name: zin.read(name) for name in zin.namelist()}

    options: list[dict] = []
    seen: set[str] = set()

    try:
        toc_entries = extract_toc_entries_from_parts(file_map)
    except Exception:
        toc_entries = []

    for entry in toc_entries:
        option = _build_section_option(entry.number, entry.title, level=entry.level, source="toc")
        if option and option["value"] not in seen:
            options.append(option)
            seen.add(option["value"])
    if options:
        return options

    document_xml = file_map.get("word/document.xml")
    styles_xml = file_map.get("word/styles.xml")
    if not document_xml or not styles_xml:
        return []

    root = etree.fromstring(document_xml)
    body = root.find("w:body", namespaces=NS)
    if body is None:
        return []

    style_outline, style_based = build_style_outline_map(styles_xml)
    style_heading_rank = build_style_heading_rank_map(styles_xml)
    body_children = list(body)
    content_children = body_children[:-1] if body_children and body_children[-1].tag == qn("w:sectPr") else body_children

    for block in content_children:
        for p in iter_paragraphs(block):
            if is_toc_paragraph(p):
                continue
            text = normalize_text(get_all_text(p))
            if not text:
                continue
            heading_depth = get_effective_heading_depth(
                p,
                style_outline,
                style_based,
                style_heading_rank,
            )
            if heading_depth is None:
                continue
            number = ""
            title = text
            match = re.match(r"^(\d+(?:[\.．]\d+)*)(?:[\.．]?\s*)(.*)$", text)
            if match:
                number = match.group(1).replace("．", ".")
                title = match.group(2)
            option = _build_section_option(number, title, level=heading_depth + 1, source="heading")
            if option and option["value"] not in seen:
                options.append(option)
                seen.add(option["value"])
    return options


def resolve_target_table_indexes(
    document_tree: etree._ElementTree,
    *,
    document_xml_path: str,
    target_chapter_ref: str = "",
    target_table_index: int | None = None,
) -> set[int] | None:
    chapter_ref = normalize_text(target_chapter_ref)
    if not chapter_ref:
        return None

    chapter_section, explicit_end_number, parsed_title = parse_chapter_section_expression(chapter_ref)
    chapter_title = parsed_title.strip()
    if not chapter_section and not chapter_title and chapter_ref:
        chapter_title = chapter_ref
    start_heading_text = chapter_title or chapter_section
    if not chapter_section and not chapter_title:
        raise ValueError("指定章節格式不正確")

    word_dir = os.path.dirname(document_xml_path)
    styles_xml_path = os.path.join(word_dir, "styles.xml")
    numbering_xml_path = os.path.join(word_dir, "numbering.xml")
    if not os.path.isfile(styles_xml_path):
        raise ValueError("Word 缺少 styles.xml，無法辨識章節範圍")

    styles_xml = open(styles_xml_path, "rb").read()
    numbering_xml = open(numbering_xml_path, "rb").read() if os.path.isfile(numbering_xml_path) else None

    style_outline, style_based = build_style_outline_map(styles_xml)
    style_heading_rank = build_style_heading_rank_map(styles_xml)
    _, style_numpr = parse_styles_numpr(styles_xml)

    root = document_tree.getroot()
    body = root.find("w:body", namespaces=NS)
    if body is None:
        raise ValueError("Word 內容格式不正確")

    body_children = list(body)
    content_children = body_children[:-1] if body_children and body_children[-1].tag == qn("w:sectPr") else body_children
    attempts = [
        {
            "start_heading_text": start_heading_text,
            "start_number": chapter_section,
            "strict_heading_number_match": True,
        },
        {
            "start_heading_text": start_heading_text,
            "start_number": chapter_section,
            "strict_heading_number_match": False,
        },
    ]
    if chapter_title:
        attempts.append({
            "start_heading_text": chapter_title,
            "start_number": "",
            "strict_heading_number_match": False,
        })

    last_error: RuntimeError | None = None
    start_idx = end_idx = None
    for attempt in attempts:
        try:
            start_idx, end_idx = find_section_range_children(
                content_children,
                start_heading_text=attempt["start_heading_text"],
                start_number=attempt["start_number"],
                style_outline=style_outline,
                style_based=style_based,
                style_numpr=style_numpr,
                style_heading_rank=style_heading_rank,
                explicit_end_number=explicit_end_number or None,
                ignore_toc=True,
                numbering_xml=numbering_xml,
                rule_based_boundary_fallback=False,
                llm_boundary_fallback=False,
                strict_heading_number_match=attempt["strict_heading_number_match"],
            )
            break
        except RuntimeError as exc:
            last_error = exc
            continue

    if start_idx is None or end_idx is None:
        raise ValueError(f"找不到指定章節：{chapter_ref}") from last_error

    scoped_tables = []
    for block in content_children[start_idx:end_idx]:
        if block.tag == qn("w:tbl"):
            scoped_tables.append(block)
        else:
            scoped_tables.extend(block.xpath(".//w:tbl", namespaces=NS))

    if target_table_index is not None:
        if target_table_index <= 0:
            raise ValueError("表格索引必須大於 0")
        if target_table_index > len(scoped_tables):
            raise ValueError(f"指定章節只有 {len(scoped_tables)} 張表，找不到第 {target_table_index} 張")
        scoped_tables = [scoped_tables[target_table_index - 1]]

    all_tables = root.xpath(".//w:tbl", namespaces=NS)
    table_index_map = {id(tbl): idx for idx, tbl in enumerate(all_tables)}
    return {table_index_map[id(tbl)] for tbl in scoped_tables if id(tbl) in table_index_map}


def expand_row_to_logical_cells(parsed_row: list[dict]) -> list[dict | None]:
    if not parsed_row:
        return []
    max_col = max(item["logical_col_end"] for item in parsed_row)
    expanded = [None] * (max_col + 1)
    for item in parsed_row:
        for col in range(item["logical_col_start"], item["logical_col_end"] + 1):
            expanded[col] = item
    return expanded


def get_row_header_options(parsed_row: list[dict]) -> list[str]:
    options: list[str] = []
    seen: set[str] = set()
    for item in parsed_row:
        text = normalize_text(item["text"])
        if not text or text in seen:
            continue
        options.append(text)
        seen.add(text)
    return options


def build_header_map_from_row(row: list[dict], header_matcher) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for header_name in AVAILABLE_HEADER_OPTIONS:
        for item in row:
            cell_text = normalize_text(item["text"])
            if cell_text and header_matcher(cell_text, header_name):
                header_map[header_name] = item["tc_idx"]
                break
    return header_map


def find_manual_header_row_and_map(
    parsed_rows: list[list[dict]],
    manual_header_mapping: dict[str, str] | None = None,
) -> tuple[int, dict[str, int], list[str]] | tuple[None, None, list[str]]:
    if not manual_header_mapping:
        return None, None, []

    normalized_mapping = {
        header_name: normalize_text(actual_text)
        for header_name, actual_text in manual_header_mapping.items()
        if header_name in AVAILABLE_HEADER_OPTIONS and normalize_text(actual_text)
    }
    if not normalized_mapping:
        return None, None, []

    missing_required = [header_name for header_name in MINIMUM_REQUIRED_HEADERS if header_name not in normalized_mapping]
    if missing_required:
        return None, None, missing_required

    if not parsed_rows:
        return None, None, list(normalized_mapping.keys())

    row = parsed_rows[0]
    header_map: dict[str, int] = {}
    for item in row:
        cell_text = normalize_text(item["text"])
        if not cell_text:
            continue
        for target_name, actual_text in normalized_mapping.items():
            if text_matches_manual_header(cell_text, actual_text):
                header_map[target_name] = item["tc_idx"]
    if all(header_name in header_map for header_name in MINIMUM_REQUIRED_HEADERS):
        return 0, header_map, []

    return None, None, list(normalized_mapping.keys())


def find_header_row_and_map(
    parsed_rows: list[list[dict]],
    required_headers: list[str] | tuple[str, ...] | None = None,
    manual_header_mapping: dict[str, str] | None = None,
) -> tuple[int, dict] | tuple[None, None]:
    manual_row_idx, manual_header_map, _ = find_manual_header_row_and_map(parsed_rows, manual_header_mapping)
    if manual_row_idx is not None and manual_header_map is not None:
        return manual_row_idx, manual_header_map

    if not parsed_rows:
        return None, None

    active_required_headers = normalize_required_headers(required_headers)
    row = parsed_rows[0]
    expanded = expand_row_to_logical_cells(row)
    texts = [x["text"] if x else "" for x in expanded]
    header_presence = {
        header_name: any(header_matches_target(t, header_name) for t in texts if t)
        for header_name in AVAILABLE_HEADER_OPTIONS
    }
    if all(header_presence.get(header_name, False) for header_name in active_required_headers):
        return 0, build_header_map_from_row(row, header_matches_target)
    return None, None


def inspect_table_headers(
    tree: etree._ElementTree,
    allowed_table_indexes: set[int] | None = None,
    manual_header_mappings: dict[int | str, dict[str, str]] | None = None,
) -> list[dict]:
    root = tree.getroot()
    tables = root.xpath(".//w:tbl", namespaces=NS)
    checks = []
    active_manual_mappings = normalize_manual_header_mappings(manual_header_mappings)
    for table_index, tbl in enumerate(tables):
        if allowed_table_indexes is not None and table_index not in allowed_table_indexes:
            continue
        parsed_rows = parse_table_rows(tbl)
        best_match = None
        header_options = get_row_header_options(parsed_rows[0]) if parsed_rows else []
        if parsed_rows:
            expanded = expand_row_to_logical_cells(parsed_rows[0])
            texts = [x["text"] if x else "" for x in expanded]
            detected_headers = [
                header_name
                for header_name in AVAILABLE_HEADER_OPTIONS
                if any(header_matches_target(t, header_name) for t in texts if t)
            ]
            if detected_headers:
                best_match = {
                    "table_index": table_index,
                    "table_label": f"表格 {table_index + 1}",
                    "header_row_index": 0,
                    "detected_headers": detected_headers,
                    "is_processable": all(header in detected_headers for header in DEFAULT_REQUIRED_HEADERS),
                    "has_optional_harmonised": "EU Harmonised Standards under MDR 2017/745 (YES/NO)" in detected_headers,
                    "has_title": "Title" in detected_headers,
                    "matched_default_count": sum(1 for header in DEFAULT_REQUIRED_HEADERS if header in detected_headers),
                }

        manual_mapping = active_manual_mappings.get(table_index, {})
        manual_mapping_summary = {
            header_name: manual_mapping.get(header_name, "")
            for header_name in AVAILABLE_HEADER_OPTIONS
        }
        manual_row_idx, manual_header_map, manual_issues = find_manual_header_row_and_map(parsed_rows, manual_mapping)
        manual_mapping_ready = manual_row_idx is not None and manual_header_map is not None
        manual_missing_headers = [
            header_name
            for header_name in MINIMUM_REQUIRED_HEADERS
            if not manual_mapping.get(header_name)
        ]

        if best_match is None:
            checks.append({
                "table_index": table_index,
                "table_label": f"表格 {table_index + 1}",
                "header_row_index": None,
                "detected_headers": [],
                "is_processable": manual_mapping_ready,
                "is_full_match": False,
                "has_optional_harmonised": False,
                "status": "manual" if manual_mapping_ready else "ignored",
                "status_label": "手動對應" if manual_mapping_ready else "未辨識",
                "message": (
                    "已套用手動欄位對應，這張表會依手動設定納入處理。"
                    if manual_mapping_ready
                    else "未找到符合預設格式的標頭資訊。若這是要更新的表格，請先完成手動對應欄位設定。"
                ),
                "header_options": header_options,
                "needs_manual_mapping": True,
                "manual_mapping": manual_mapping_summary,
                "manual_mapping_ready": manual_mapping_ready,
                "manual_missing_headers": manual_missing_headers,
                "manual_mapping_issues": manual_issues,
            })
            continue

        missing_default = [header for header in DEFAULT_REQUIRED_HEADERS if header not in best_match["detected_headers"]]
        is_full_match = all(header in best_match["detected_headers"] for header in DEFAULT_REQUIRED_HEADERS)

        if is_full_match:
            best_match["status"] = "full"
            best_match["status_label"] = "完整格式"
            best_match["message"] = "已找到 Standards、Issued Year、EU Harmonised 與 Title 欄位。"
        elif manual_mapping_ready:
            best_match["status"] = "manual"
            best_match["status_label"] = "手動對應"
            best_match["message"] = "此表格不符合預設四欄格式，但已完成手動對應，可進行更新。"
        else:
            best_match["status"] = "partial"
            best_match["status_label"] = "需手動對應"
            best_match["message"] = f"與預設四欄格式不一致，缺少或無法辨識：{'、'.join(missing_default)}。請先完成手動對應欄位設定。"
        best_match["header_options"] = header_options
        best_match["is_processable"] = is_full_match or manual_mapping_ready
        best_match["is_full_match"] = is_full_match
        best_match["harmonised_update_supported"] = (
            "EU Harmonised Standards under MDR 2017/745 (YES/NO)" in best_match.get("detected_headers", [])
            or bool((manual_mapping_summary if manual_mapping_ready else {}).get("EU Harmonised Standards under MDR 2017/745 (YES/NO)"))
        )
        if best_match["is_processable"] and not best_match["harmonised_update_supported"]:
            best_match["message"] = f"{best_match['message']} 此表未包含 EU Harmonised Standards under MDR 2017/745 (YES/NO) 欄位，更新時會略過該欄。"
        best_match["needs_manual_mapping"] = (not is_full_match) or bool(manual_mapping)
        best_match["manual_mapping"] = manual_mapping_summary
        best_match["manual_mapping_ready"] = manual_mapping_ready
        best_match["manual_missing_headers"] = manual_missing_headers
        best_match["manual_mapping_issues"] = manual_issues
        checks.append(best_match)
    return checks


def get_logical_col(header_map: dict, target_name: str) -> int | None:
    return header_map.get(target_name)


def parse_word_tables_for_update(
    document_xml_path: str,
    required_headers: list[str] | tuple[str, ...] | None = None,
    allowed_table_indexes: set[int] | None = None,
    manual_header_mappings: dict[int | str, dict[str, str]] | None = None,
) -> tuple[etree._ElementTree, list[dict]]:
    tree = etree.parse(document_xml_path)
    root = tree.getroot()
    tables = root.xpath(".//w:tbl", namespaces=NS)
    all_records = []
    active_required_headers = normalize_required_headers(required_headers)
    active_manual_mappings = normalize_manual_header_mappings(manual_header_mappings)
    for table_index, tbl in enumerate(tables):
        if allowed_table_indexes is not None and table_index not in allowed_table_indexes:
            continue
        parsed_rows = parse_table_rows(tbl)
        header_row_idx, header_map = find_header_row_and_map(
            parsed_rows,
            active_required_headers,
            active_manual_mappings.get(table_index),
        )
        if header_row_idx is None:
            continue
        standards_col = get_logical_col(header_map, "Standards")
        issued_year_col = get_logical_col(header_map, "Issued Year")
        harmonised_col = get_logical_col(header_map, "EU Harmonised Standards under MDR 2017/745 (YES/NO)")
        title_col = get_logical_col(header_map, "Title")
        if standards_col is None or issued_year_col is None:
            continue
        current_category = ""
        for row_idx in range(header_row_idx + 1, len(parsed_rows)):
            row_items = parsed_rows[row_idx]

            def get_text_at(tc_idx: int | None) -> str:
                if tc_idx is None or tc_idx >= len(row_items):
                    return ""
                return normalize_text(row_items[tc_idx]["text"])

            def get_tc_at(tc_idx: int | None):
                if tc_idx is None or tc_idx >= len(row_items):
                    return None
                return row_items[tc_idx]["tc"]

            standards_text = get_text_at(standards_col)
            issued_year_text = get_text_at(issued_year_col)
            harmonised_text = get_text_at(harmonised_col)
            title_text = get_text_at(title_col)
            nonempty_texts = [normalize_text(item["text"]) for item in row_items if normalize_text(item["text"])]

            is_single_cell_category = len(row_items) == 1 and len(nonempty_texts) == 1
            if is_single_cell_category:
                current_category = nonempty_texts[0]
                continue

            if standards_text == "" and issued_year_text == "" and harmonised_text == "" and title_text == "":
                if nonempty_texts:
                    current_category = nonempty_texts[0]
                continue

            if standards_text:
                all_records.append({
                    "table_index": table_index,
                    "row_index": row_idx,
                    "category": current_category,
                    "standards": standards_text,
                    "issued_year": issued_year_text,
                    "harmonised": harmonised_text,
                    "title": title_text,
                    "standards_tc": get_tc_at(standards_col),
                    "issued_year_tc": get_tc_at(issued_year_col),
                    "harmonised_tc": get_tc_at(harmonised_col),
                    "title_tc": get_tc_at(title_col),
                })
    return tree, all_records


def build_preview_tables(
    tree: etree._ElementTree,
    row_reference_map: dict,
    required_headers: list[str] | tuple[str, ...] | None = None,
    allowed_table_indexes: set[int] | None = None,
    manual_header_mappings: dict[int | str, dict[str, str]] | None = None,
) -> tuple[list[dict], dict]:
    root = tree.getroot()
    tables = root.xpath(".//w:tbl", namespaces=NS)
    preview_tables = []
    reference_payload: dict[str, dict] = {}
    table_number = 0
    active_required_headers = normalize_required_headers(required_headers)
    active_manual_mappings = normalize_manual_header_mappings(manual_header_mappings)

    for table_index, tbl in enumerate(tables):
        if allowed_table_indexes is not None and table_index not in allowed_table_indexes:
            continue
        parsed_rows = parse_table_rows(tbl)
        header_row_idx, header_map = find_header_row_and_map(
            parsed_rows,
            active_required_headers,
            active_manual_mappings.get(table_index),
        )
        if header_row_idx is None:
            continue
        standards_col = get_logical_col(header_map, "Standards")
        issued_year_col = get_logical_col(header_map, "Issued Year")
        harmonised_col = get_logical_col(header_map, "EU Harmonised Standards under MDR 2017/745 (YES/NO)")
        title_col = get_logical_col(header_map, "Title")
        header_row_items = parsed_rows[header_row_idx]
        header_labels = {
            item["tc_idx"]: normalize_text(item["text"])
            for item in header_row_items
        }
        fixed_columns = [
            ("Standards", standards_col),
            ("Issued Year", issued_year_col),
            ("EU Harmonised Standards under MDR 2017/745 (YES/NO)", harmonised_col),
            ("Title", title_col),
        ]
        display_columns = [(label, tc_idx) for label, tc_idx in fixed_columns if tc_idx is not None]
        table_number += 1
        rows_data = []

        for row_idx, row in enumerate(parsed_rows):
            def get_item(tc_idx: int | None):
                if tc_idx is None or tc_idx >= len(row):
                    return None
                return row[tc_idx]

            row_classes = []
            if row_idx == header_row_idx:
                row_classes.append("is-header")
            elif row_idx > header_row_idx:
                nonempty_texts = [normalize_text(item["text"]) for item in row if normalize_text(item["text"])]
                is_single_cell_category = len(row) == 1 and len(nonempty_texts) == 1
                standards_item = get_item(standards_col)
                issued_year_item = get_item(issued_year_col)
                harmonised_item = get_item(harmonised_col)
                title_item = get_item(title_col)
                standards_text = normalize_text(standards_item["text"]) if standards_item else ""
                issued_year_text = normalize_text(issued_year_item["text"]) if issued_year_item else ""
                harmonised_text = normalize_text(harmonised_item["text"]) if harmonised_item else ""
                title_text = normalize_text(title_item["text"]) if title_item else ""
                if is_single_cell_category or (
                    standards_text == ""
                    and issued_year_text == ""
                    and harmonised_text == ""
                    and title_text == ""
                    and nonempty_texts
                ):
                    row_classes.append("is-category")

            row_meta = row_reference_map.get((table_index, row_idx))
            if row_meta:
                row_classes.append(f"status-{row_meta.get('status', '').lower().replace(':', '-')}")

            cells = []
            if row_idx == header_row_idx:
                for label, tc_idx in display_columns:
                    item = get_item(tc_idx)
                    cells.append({
                        "tag": "th",
                        "colspan": 1,
                        "content_html": format_cell_runs_as_html(item["tc"]) if item is not None else html.escape(label),
                        "reference_key": None,
                        "header_text": label,
                    })
            elif "is-category" in row_classes:
                category_item = next((item for item in row if normalize_text(item["text"])), None)
                cells.append({
                    "tag": "td",
                    "colspan": len(display_columns) or 1,
                    "content_html": format_cell_runs_as_html(category_item["tc"]) if category_item is not None else "",
                    "reference_key": None,
                    "header_text": "",
                })
            else:
                for label, tc_idx in display_columns:
                    item = get_item(tc_idx)
                    reference_key = None
                    if row_meta and label == "Standards":
                        reference_key = f"{row_meta['row_key']}:standards"
                        reference_payload[reference_key] = {**row_meta, "field_label": "Standards"}
                    elif row_meta and label == "Issued Year":
                        reference_key = f"{row_meta['row_key']}:issued_year"
                        reference_payload[reference_key] = {**row_meta, "field_label": "Issued Year"}
                    elif row_meta and label == "EU Harmonised Standards under MDR 2017/745 (YES/NO)":
                        reference_key = f"{row_meta['row_key']}:harmonised"
                        reference_payload[reference_key] = {**row_meta, "field_label": "EU Harmonised Standards under MDR 2017/745 (YES/NO)"}
                    elif row_meta and label == "Title":
                        reference_key = f"{row_meta['row_key']}:title"
                        reference_payload[reference_key] = {**row_meta, "field_label": "Title"}
                    cells.append({
                        "tag": "td",
                        "colspan": 1,
                        "content_html": format_cell_runs_as_html(item["tc"]) if item is not None else "",
                        "reference_key": reference_key,
                        "header_text": header_labels.get(tc_idx, label),
                    })

            rows_data.append({
                "classes": row_classes,
                "cells": cells,
                "row_key": row_meta.get("row_key", "") if row_meta else "",
            })

        preview_tables.append({
            "title": f"表格 {table_number}",
            "rows": rows_data,
        })

    return preview_tables, reference_payload


def inspect_document_tables(
    word_path: str,
    *,
    target_chapter_ref: str = "",
    target_table_index: int | None = None,
    manual_header_mappings: dict[int | str, dict[str, str]] | None = None,
) -> dict:
    normalized_manual_header_mappings = normalize_manual_header_mappings(manual_header_mappings)
    with tempfile.TemporaryDirectory() as tmpdir:
        unzip_docx(word_path, tmpdir)
        document_xml_path = os.path.join(tmpdir, "word", "document.xml")
        tree = etree.parse(document_xml_path)
        allowed_table_indexes = resolve_target_table_indexes(
            tree,
            document_xml_path=document_xml_path,
            target_chapter_ref=target_chapter_ref,
            target_table_index=target_table_index,
        )
        table_checks = inspect_table_headers(
            tree,
            allowed_table_indexes=allowed_table_indexes,
            manual_header_mappings=normalized_manual_header_mappings,
        )
        return {
            "table_checks": table_checks,
            "target_chapter_ref": normalize_text(target_chapter_ref),
            "target_table_index": target_table_index,
            "scope_table_count": len(allowed_table_indexes or []),
            "inspection_only": True,
            "manual_header_mappings": {
                str(table_index): mapping
                for table_index, mapping in normalized_manual_header_mappings.items()
            },
        }


def process_document(
    word_path: str,
    excel_path: str,
    harmonised_reference_path: str | None = None,
    regulation_reference_path: str | None = None,
    override_map: dict | None = None,
    output_path: str | None = None,
    iso_priority: list[str] | tuple[str, ...] | None = None,
    enabled_standard_levels: list[str] | tuple[str, ...] | None = None,
    prefer_latest_en_variants: bool = DEFAULT_PREFER_LATEST_EN_VARIANTS,
    required_headers: list[str] | tuple[str, ...] | None = None,
    target_chapter_ref: str = "",
    target_table_index: int | None = None,
    manual_header_mappings: dict[int | str, dict[str, str]] | None = None,
) -> dict:
    override_map = override_map or {}
    normalized_iso_priority = normalize_iso_priority(iso_priority)
    normalized_enabled_levels = normalize_enabled_standard_levels(enabled_standard_levels)
    normalized_required_headers = normalize_required_headers(required_headers)
    normalized_manual_header_mappings = normalize_manual_header_mappings(manual_header_mappings)
    excel_index = load_excel_index(excel_path)
    harmonised_reference_index = load_harmonised_reference_index(harmonised_reference_path)
    regulation_reference_index = load_regulation_reference_index(regulation_reference_path)
    with tempfile.TemporaryDirectory() as tmpdir:
        unzip_docx(word_path, tmpdir)
        document_xml_path = os.path.join(tmpdir, "word", "document.xml")
        tree = etree.parse(document_xml_path)
        allowed_table_indexes = resolve_target_table_indexes(
            tree,
            document_xml_path=document_xml_path,
            target_chapter_ref=target_chapter_ref,
            target_table_index=target_table_index,
        )
        table_checks = inspect_table_headers(
            tree,
            allowed_table_indexes=allowed_table_indexes,
            manual_header_mappings=normalized_manual_header_mappings,
        )
        tree, records = parse_word_tables_for_update(
            document_xml_path,
            normalized_required_headers,
            allowed_table_indexes=allowed_table_indexes,
            manual_header_mappings=normalized_manual_header_mappings,
        )
        report = []
        updated_count = 0
        row_reference_map = {}

        for rec in records:
            row_key = make_row_key(rec["table_index"], rec["row_index"])
            standards = rec["standards"]
            word_year_text = normalize_text(rec["issued_year"])
            word_harmonised_text = normalize_text(rec["harmonised"])
            word_title_text = normalize_text(rec["title"])
            match_info = None
            if is_regulation_lookup_target(standards):
                match_info = find_latest_year_from_regulation_reference(
                    standards,
                    regulation_reference_index,
                )
            if not match_info:
                match_info = find_latest_year_from_excel(
                    standards,
                    excel_index,
                    normalized_iso_priority,
                    normalized_enabled_levels,
                    harmonised_reference_index=harmonised_reference_index,
                    prefer_latest_en_variants=prefer_latest_en_variants,
                )
            if match_info:
                match_info = apply_candidate_override(match_info, override_map.get(row_key))

            if not match_info:
                if rec["standards_tc"] is not None:
                    rebuild_cell_with_single_color(rec["standards_tc"], standards, BLUE_COLOR)
                row_reference_map[(rec["table_index"], rec["row_index"])] = {
                    "row_key": row_key,
                    "status": "NOT_FOUND",
                    "sheet_name": "",
                    "excel_col_letter": "",
                    "excel_row_index": "",
                    "matched_standard_no": "",
                    "matched_display_standard_no": "",
                    "matched_harmonised": "",
                    "matched_title": "",
                    "excel_year": "",
                    "word_standard": standards,
                    "word_year": word_year_text,
                    "word_harmonised": word_harmonised_text,
                    "word_title": word_title_text,
                    "all_candidates": [],
                    "selected_candidate_id": "",
                    "auto_selected_candidate_id": "",
                }
                report.append({
                    "status": "NOT_FOUND",
                    "category": rec["category"],
                    "standards": standards,
                    "word_year": word_year_text,
                    "excel_year": "",
                    "sheet_name": "",
                    "excel_col_letter": "",
                    "excel_row_index": "",
                    "matched_standard_no": "",
                })
                continue

            apply_year_comparison = bool(match_info.get("apply_year_comparison"))
            latest_year = "" if match_info.get("latest_year") in {None, ""} else str(match_info["latest_year"])
            matched_standard_no = match_info["matched_standard_no"]
            matched_display_standard_no = match_info["matched_display_standard_no"]
            matched_harmonised = normalize_text(match_info.get("matched_harmonised", ""))
            matched_title = build_title_with_amendment(match_info.get("matched_title", ""), matched_standard_no)
            standards_needs_update = normalize_key_for_search(standards) != normalize_key_for_search(matched_display_standard_no)
            year_needs_update = bool(latest_year) and word_year_text != latest_year
            harmonised_needs_update = normalize_key_for_search(word_harmonised_text) != normalize_key_for_search(matched_harmonised)
            title_needs_update = normalize_key_for_search(word_title_text) != normalize_key_for_search(matched_title)

            if standards_needs_update and rec["standards_tc"] is not None:
                rebuild_cell_with_segments(rec["standards_tc"], build_diff_segments(standards, matched_display_standard_no))
            if year_needs_update and rec["issued_year_tc"] is not None:
                rebuild_cell_with_segments(rec["issued_year_tc"], build_year_segments(word_year_text, latest_year))
            if harmonised_needs_update and rec["harmonised_tc"] is not None:
                rebuild_cell_with_segments(rec["harmonised_tc"], build_preserve_original_segments(word_harmonised_text, matched_harmonised))
            if title_needs_update and rec["title_tc"] is not None:
                rebuild_cell_with_segments(rec["title_tc"], build_diff_segments(word_title_text, matched_title))

            row_updated = standards_needs_update or year_needs_update or harmonised_needs_update or title_needs_update
            if row_updated:
                updated_count += 1

            status = "UPDATED" if row_updated else "SAME_NO_UPDATE"
            row_reference_map[(rec["table_index"], rec["row_index"])] = {
                "row_key": row_key,
                "status": status,
                "sheet_name": match_info["sheet_name"],
                "excel_col_letter": match_info["excel_col_letter"],
                "excel_row_index": match_info["excel_row_index"],
                "matched_standard_no": matched_standard_no,
                "matched_display_standard_no": matched_display_standard_no,
                "matched_harmonised": matched_harmonised,
                "matched_title": matched_title,
                "excel_year": latest_year,
                "word_standard": standards,
                "word_year": word_year_text,
                "word_harmonised": word_harmonised_text,
                "word_title": word_title_text,
                "all_candidates": match_info.get("all_candidates", []),
                "selected_candidate_id": match_info.get("selected_candidate_id", ""),
                "auto_selected_candidate_id": match_info.get("auto_selected_candidate_id", ""),
            }
            report.append({
                "status": status,
                "category": rec["category"],
                "standards": standards,
                "word_year": word_year_text,
                "excel_year": latest_year,
                "sheet_name": match_info["sheet_name"],
                "excel_col_letter": match_info["excel_col_letter"],
                "excel_row_index": match_info["excel_row_index"],
                "matched_standard_no": matched_standard_no,
            })

        tree.write(document_xml_path, xml_declaration=True, encoding="UTF-8", standalone="yes")
        preview_tables, reference_payload = build_preview_tables(
            tree,
            row_reference_map,
            normalized_required_headers,
            allowed_table_indexes=allowed_table_indexes,
            manual_header_mappings=normalized_manual_header_mappings,
        )
        if output_path:
            zip_to_docx(tmpdir, output_path)
        return {
            "report": report,
            "updated_count": updated_count,
            "preview_tables": preview_tables,
            "reference_payload": reference_payload,
            "harmonised_reference_path": normalize_text(harmonised_reference_path or ""),
            "regulation_reference_path": normalize_text(regulation_reference_path or ""),
            "iso_priority": list(normalized_iso_priority),
            "enabled_standard_levels": list(normalized_enabled_levels),
            "prefer_latest_en_variants": prefer_latest_en_variants,
            "required_headers": list(normalized_required_headers),
            "manual_header_mappings": {
                str(table_index): mapping
                for table_index, mapping in normalized_manual_header_mappings.items()
            },
            "table_checks": table_checks,
            "target_chapter_ref": normalize_text(target_chapter_ref),
            "target_table_index": target_table_index,
            "scope_table_count": len(allowed_table_indexes or []),
            "inspection_only": False,
        }
