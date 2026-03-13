import os
import re
import tempfile
import uuid
import shutil
import json
import zipfile
from collections import defaultdict
from typing import Dict, List, Tuple, Any

from docx import Document as DocxDocument
from spire.doc import Document, FileFormat

from .Edit_Word import (
    renumber_figures_tables_file,
    insert_numbered_heading,
    insert_roman_heading,
    insert_bulleted_heading,
)
from .Extract_AllFile_to_FinalWord import (
    extract_word_all_content,
    extract_word_chapter,
    extract_specific_figure_from_word,
    extract_specific_table_from_word,
    apply_basic_style,
    remove_hidden_runs,
    hide_paragraphs_with_text,
)
from .file_copier import copy_files
from .docx_merger import merge_word_docs
from .template_manager import parse_template_paragraphs, render_template_with_mappings
from .workflow import run_workflow
from app.services.flow_service import (
    DEFAULT_APPLY_FORMATTING,
    DEFAULT_DOCUMENT_FORMAT_KEY,
    DEFAULT_LINE_SPACING,
    DOCUMENT_FORMAT_PRESETS,
    SKIP_DOCX_CLEANUP,
    collect_titles_to_hide,
)


def _split_rel_parts(path: str) -> list[str]:
    return [part for part in re.split(r"[\\/]+", str(path or "").strip()) if part]


def _find_files(base: str, filename: str) -> list[str]:
    """Search *base* recursively for *filename* ignoring case."""
    target = filename.lower()
    matches: list[str] = []
    for root, _dirs, files in os.walk(base):
        for fn in files:
            if fn.lower() == target:
                matches.append(os.path.join(root, fn))
    return matches


def _find_file(base: str, filename: str) -> str | None:
    matches = _find_files(base, filename)
    return matches[0] if matches else None


def _find_directory(base: str, path: str) -> str | None:
    """Locate a directory relative to *base* ignoring case."""
    parts = _split_rel_parts(path)
    current = base
    for part in parts:
        match = None
        for name in os.listdir(current):
            candidate = os.path.join(current, name)
            if os.path.isdir(candidate) and name.lower() == part.lower():
                match = candidate
                break
        if match is None:
            return None
        current = match
    return current


def _resolve_relative_file(base: str, path: str) -> str | None:
    parts = _split_rel_parts(path)
    if not parts:
        return None
    current = base
    for part in parts[:-1]:
        match = None
        for name in os.listdir(current):
            candidate = os.path.join(current, name)
            if os.path.isdir(candidate) and name.lower() == part.lower():
                match = candidate
                break
        if match is None:
            return None
        current = match
    target_name = parts[-1].lower()
    for name in os.listdir(current):
        candidate = os.path.join(current, name)
        if os.path.isfile(candidate) and name.lower() == target_name:
            return candidate
    return None


def _resolve_input_file(base: str, name: str) -> tuple[str | None, str]:
    """Resolve *name* to a file path.

    If *name* includes an extension and contains path separators, it is treated
    as a relative file path under *base* and resolved case-insensitively.
    If *name* includes an extension without path separators, it is treated as a
    filename and searched within *base*.
    If it has no extension, it is treated as a directory and the first document
    file inside that directory is returned.
    """

    raw_name = str(name or "").strip()
    if not raw_name:
        return None, "empty input name"

    if "." in os.path.basename(raw_name):
        if "/" in raw_name or "\\" in raw_name:
            resolved = _resolve_relative_file(base, raw_name)
            if resolved:
                return resolved, ""
            return None, f"file not found: {raw_name}"
        matches = _find_files(base, raw_name)
        if not matches:
            return None, f"file not found: {raw_name}"
        if len(matches) > 1:
            rel_matches = sorted(os.path.relpath(p, base).replace("\\", "/") for p in matches)
            joined = "; ".join(rel_matches)
            return None, f"找到多個名稱相同檔案 {raw_name}: {joined}"
        return matches[0], ""

    dir_path = _find_directory(base, raw_name)
    if not dir_path:
        return None, f"directory not found: {raw_name}"
    for fn in os.listdir(dir_path):
        if fn.lower().endswith((".docx", ".doc")):
            return os.path.join(dir_path, fn), ""
    return None, f"no Word file found in directory: {raw_name}"




def _normalize_match(text: str) -> str:
    cleaned = (text or "").strip().lower()
    cleaned = re.sub(r"\s+", "", cleaned)
    return cleaned


def _normalize_output_rel_path(path: str) -> str:
    """Treat both Windows and POSIX separators as directory delimiters."""
    raw = (path or "").strip()
    if not raw:
        return ""
    parts = [part for part in re.split(r"[\\/]+", raw) if part]
    if not parts:
        return ""
    return os.path.join(*parts)


def _find_header_row(ws, header_names: List[str], max_scan: int = 10) -> int | None:
    max_row = ws.max_row or 0
    scan = min(max_row, max_scan)
    for row_idx in range(1, scan + 1):
        row_vals = [
            str(c.value).strip() if c is not None and c.value is not None else ""
            for c in ws[row_idx]
        ]
        if all(h in row_vals for h in header_names):
            return row_idx
    return None


def _build_template_index_map(parsed: List[Dict[str, Any]]) -> Tuple[Dict[str, int], int | None]:
    index_map: Dict[str, int] = {}
    last_idx = None
    for item in parsed:
        try:
            idx = int(item.get("index"))
        except Exception:
            continue
        display = (item.get("display") or "").strip()
        text_val = (item.get("text") or "").strip()
        key = _normalize_match(f"{display} {text_val}".strip())
        if key and key not in index_map:
            index_map[key] = idx
        last_idx = idx
    return index_map, last_idx


def insert_title(section, title: str):
    """Insert *title* into *section* with appropriate heading style.

    - Titles beginning with Roman numerals (e.g. ``"I."``, ``"II."``) use
      :func:`insert_roman_heading`.
    - Titles beginning with a ``"⚫"`` bullet use :func:`insert_bulleted_heading`.
    - All other titles use :func:`insert_numbered_heading`.
    """

    if not title:
        return None

    # Strip leading chapter numbers like "6.4.2" from the title
    title = re.sub(r"^[0-9]+(?:\.[0-9]+)*\s*", "", title)

    roman_match = re.match(r"^[IVXLCDM]+\.\s*(.*)", title)
    if roman_match:
        text = roman_match.group(1).strip() or title
        return insert_roman_heading(section, text, level=0, bold=True, font_size=12)

    if title.startswith("⚫"):
        text = title.lstrip("⚫").strip()
        return insert_bulleted_heading(section, text, level=0, bullet_char='·', bold=True, font_size=12)

    return insert_numbered_heading(section, title, level=0, bold=True, font_size=12)

def process_mapping_excel(
    mapping_path: str,
    task_files_dir: str,
    output_dir: str,
    log_dir: str | None = None,
    validate_only: bool = False,
    validate_extract_only: bool = False,
) -> Dict[str, List[str]]:
    """Process mapping Excel file and generate documents or copy files.

    New format columns:
        A: Source file or text
        B: Section/Operation
        C: Type
        D: Include title/caption
        E: Output path
        F: Output filename
        G: Template file
        H: Insert paragraph
    Returns a dict with keys:
        logs: list of messages
        outputs: list of generated docx paths
    """
    if validate_only and validate_extract_only:
        raise ValueError("validate_only and validate_extract_only cannot both be True")

    logs: List[str] = []
    outputs: List[str] = []
    row_errors: Dict[int, List[str]] = defaultdict(list)
    os.makedirs(output_dir, exist_ok=True)

    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required to process mapping files") from e

    wb = load_workbook(mapping_path)
    ws = wb.worksheets[0]

    header_aliases = {
        "source": ["檔案名稱/資料夾名稱/文字內容", "來源檔案"],
        "operation": ["擷取段落", "擷取段落"],
        "out_path": ["檔案路徑", "輸出路徑"],
        "out_name": ["檔案名稱", "輸出檔案名稱"],
        "template": ["模板文件"],
        "insert": ["插入段落名稱/目的資料夾名稱", "插入段落"],
    }
    optional_header_aliases = {
        "item_type": ["類型", "Type", "擷取類型"],
        "include_title": ["包含標題", "是否包含標題", "顯示標題", "Include Title", "Include Caption"],
    }
    header_row = None
    max_row = ws.max_row or 0
    scan = min(max_row, 10)
    for row_idx in range(1, scan + 1):
        row_vals = [
            str(c.value).strip() if c is not None and c.value is not None else ""
            for c in ws[row_idx]
        ]
        if all(any(alias in row_vals for alias in aliases) for aliases in header_aliases.values()):
            header_row = row_idx
            break


    if header_row is None:
        # Fallback to legacy format
        start_row = 3 if ws.max_row and ws.max_row >= 3 else 2
        docs: Dict[str, Tuple[Document, Any]] = {}
        hidden_titles: Dict[str, List[str]] = defaultdict(list)
        if validate_only or validate_extract_only:
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                raw_out, _raw_title, raw_folder, raw_input, raw_instruction = row[:5]
                out_name = str(raw_out).strip() if raw_out else ""
                folder = str(raw_folder).strip() if raw_folder else ""
                input_name = str(raw_input).strip() if raw_input else ""
                instruction = str(raw_instruction).strip() if raw_instruction else ""
                if not instruction:
                    continue

                base_dir = task_files_dir
                if folder:
                    found_dir = _find_directory(task_files_dir, folder)
                    if not found_dir:
                        logs.append(f"ERROR: {out_name or '?'} folder not found {folder}")
                        continue
                    base_dir = found_dir

                is_all = instruction.lower() == "all"
                chapter_match = re.match(r"^([0-9]+(?:\.[0-9]+)*)(?:.*)", instruction)
                if is_all or chapter_match:
                    if not input_name:
                        logs.append(f"ERROR: {out_name or '?'} missing source filename")
                        continue
                    infile, resolve_error = _resolve_input_file(base_dir, input_name)
                    if not infile:
                        logs.append(f"ERROR: {out_name or '?'} {resolve_error}")
                        continue

            log_file = None
            if logs:
                target_log_dir = log_dir or output_dir
                os.makedirs(target_log_dir, exist_ok=True)
                log_filename = f"mapping_log_{uuid.uuid4().hex[:8]}.json"
                log_path = os.path.join(target_log_dir, log_filename)
                with open(log_path, "w", encoding="utf-8") as f:
                    json.dump({"messages": logs, "runs": []}, f, ensure_ascii=False, indent=2)
                log_file = log_filename
            return {"logs": logs, "outputs": [], "log_file": log_file}

        for row in ws.iter_rows(min_row=start_row, values_only=True):
            raw_out, raw_title, raw_folder, raw_input, raw_instruction = row[:5]
            out_name = str(raw_out).strip() if raw_out else ""
            title = str(raw_title).strip() if raw_title else ""
            folder = str(raw_folder).strip() if raw_folder else ""
            input_name = str(raw_input).strip() if raw_input else ""
            instruction = str(raw_instruction).strip() if raw_instruction else ""
            if not instruction:
                continue

            base_dir = task_files_dir
            if folder:
                found_dir = _find_directory(task_files_dir, folder)
                if not found_dir:
                    logs.append(f"{out_name or '?'}: folder not found {folder}")
                    continue
                base_dir = found_dir

            is_all = instruction.lower() == "all"
            chapter_match = re.match(r"^([0-9]+(?:\.[0-9]+)*)(?:.*)", instruction)

            if is_all or chapter_match:
                if not input_name:
                    logs.append(f"{out_name or '?'}: missing source filename")
                    continue
                infile, resolve_error = _resolve_input_file(base_dir, input_name)
                if not infile:
                    logs.append(f"{out_name or '?'}: {resolve_error}")
                    continue

                doc, section = docs.get(out_name, (None, None))
                if doc is None:
                    doc = Document()
                    section = doc.AddSection()
                    docs[out_name] = (doc, section)

                insert_title(section, title)

                if is_all:
                    extract_word_all_content(infile, output_doc=doc, section=section)
                    logs.append(f"Extract {input_name} (all content)")
                else:
                    parsed_chapter, parsed_title, _parsed_sub = _parse_chapter_parts(instruction)
                    chapter = parsed_chapter or chapter_match.group(1)
                    title_inline = parsed_title or ""
                    if "," in instruction and not title_inline:
                        _prefix, after = instruction.split(",", 1)
                        title_inline = after.strip()
                    if title_inline:
                        result = extract_word_chapter(
                            infile,
                            chapter,
                            use_chapter_title=True,
                            target_chapter_title=title_inline,
                            output_doc=doc,
                            section=section,
                        )
                        if isinstance(result, dict):
                            for captured in result.get("captured_titles", []):
                                if not isinstance(captured, str):
                                    continue
                                trimmed = captured.strip()
                                if trimmed and trimmed not in hidden_titles[out_name]:
                                    hidden_titles[out_name].append(trimmed)
                        logs.append(f"Extract {input_name} (chapter {chapter}, title {title_inline})")
                    else:
                        result = extract_word_chapter(
                            infile,
                            chapter,
                            output_doc=doc,
                            section=section,
                        )
                        if isinstance(result, dict):
                            for captured in result.get("captured_titles", []):
                                if not isinstance(captured, str):
                                    continue
                                trimmed = captured.strip()
                                if trimmed and trimmed not in hidden_titles[out_name]:
                                    hidden_titles[out_name].append(trimmed)
                        logs.append(f"Extract {input_name} (chapter {chapter})")
            else:
                dest = os.path.join(task_files_dir, out_name or "output")
                if title:
                    dest = os.path.join(dest, title)

                search_root = base_dir
                if input_name:
                    if "." in os.path.basename(input_name):
                        found, _resolve_error = _resolve_input_file(base_dir, input_name)
                        if found:
                            search_root = os.path.dirname(found)
                    else:
                        dir_path = _find_directory(base_dir, input_name)
                        if dir_path:
                            search_root = dir_path

                keywords = [k.strip() for k in re.split(r"[,\u3001，]+", instruction) if k.strip()]
                try:
                    copied = copy_files(search_root, dest, keywords)
                    kw_display = ", ".join(keywords)
                    logs.append(
                        f"Copied {len(copied)} files to {os.path.relpath(dest, task_files_dir)}"
                        f" (keywords {kw_display})"
                    )
                except Exception as e:
                    logs.append(f"Copy failed: {e}")

        os.makedirs(output_dir, exist_ok=True)
        for name, (doc, _section) in docs.items():
            out_path = os.path.join(output_dir, f"{name}.docx")
            doc.SaveToFile(out_path, FileFormat.Docx)
            doc.Close()
            titles = hidden_titles.get(name, [])
            remove_hidden_runs(out_path, preserve_texts=titles)
            renumber_figures_tables_file(out_path)
            apply_basic_style(out_path)
            hide_paragraphs_with_text(out_path, titles)
            outputs.append(out_path)
        return {"logs": logs, "outputs": outputs}

    # New format processing
    header_vals = [str(c.value).strip() if c is not None and c.value is not None else "" for c in ws[header_row]]
    col_idx = {}
    for key, aliases in header_aliases.items():
        for alias in aliases:
            if alias in header_vals:
                col_idx[key] = header_vals.index(alias)
                break
    for key, aliases in optional_header_aliases.items():
        for alias in aliases:
            if alias in header_vals:
                col_idx[key] = header_vals.index(alias)
                break
    parsed_cache: Dict[str, Tuple[List[Dict[str, Any]], Dict[str, int], int | None]] = {}
    groups: Dict[Tuple[str, str | None], Dict[str, Any]] = {}
    run_logs: List[Dict[str, Any]] = []
    output_template_map: Dict[str, str | None] = {}

    def _log(
        level: str,
        message: str,
        row_num: int | None = None,
        action: str | None = None,
        detail: str | None = None,
    ) -> None:
        row_tag = f"(Row {row_num}) " if row_num else ""
        if level.lower() == "error" and action:
            if row_num:
                row_errors[row_num].append(message)
            if detail:
                logs.append(f"{level.upper()}: {row_tag}{action} :: {detail} :: {message}")
            else:
                logs.append(f"{level.upper()}: {row_tag}{action}: {message}")
        else:
            prefix = f"Row {row_num}: " if row_num else ""
            logs.append(f"{level.upper()}: {prefix}{message}")

    def _normalize_item_type(raw_type: str) -> str:
        text = (raw_type or "").strip().lower()
        if text in {"figure", "fig", "圖片", "图"}:
            return "figure"
        if text in {"table", "表格", "表"}:
            return "table"
        if text in {"add text", "text", "文字", "加入文字", "新增文字"}:
            return "add_text"
        if text in {"pdf image", "pdf images", "pdfimage", "pdf_img", "pdf圖片", "pdf图", "pdf 圖片", "pdf 图片"}:
            return "pdf_image"
        return ""

    def _parse_include_title(raw_value: str) -> tuple[bool, str]:
        text = (raw_value or "").strip()
        if not text:
            return True, ""
        normalized = re.sub(r"\s+", "", text).lower()
        if normalized in {
            "1",
            "true",
            "yes",
            "y",
            "on",
            "include",
            "show",
            "keep",
            "包含",
            "顯示",
            "保留",
            "是",
            "要",
        }:
            return True, ""
        if normalized in {
            "0",
            "false",
            "no",
            "n",
            "off",
            "exclude",
            "hide",
            "omit",
            "remove",
            "不包含",
            "不顯示",
            "隱藏",
            "否",
            "不要",
            "移除",
        }:
            return False, ""
        return True, f"包含標題欄位值無效: {text}"

    def _guess_action(instruction: str, item_type: str = "") -> str:
        if item_type == "add_text":
            return "Append text"
        if item_type == "figure":
            return "Extract figure"
        if item_type == "table":
            return "Extract table"
        if item_type == "pdf_image":
            return "Extract PDF images"
        ins = (instruction or "").strip()
        if not ins:
            return "Mapping"
        low = ins.lower()
        if low == "add text":
            return "Append text"
        label_match = re.search(r"\b(Table|Figure)\b", ins, re.IGNORECASE)
        if label_match:
            return "Extract table" if label_match.group(1).lower() == "table" else "Extract figure"
        if low == "all":
            return "Extract all"
        if re.match(r"^\d+(?:\.\d+)*", ins):
            return "Extract chapter"
        return "Mapping"

    def _parse_chapter_parts(text: str) -> tuple[str, str, str]:
        chapter = ""
        title = ""
        subheading = ""
        if not text:
            return chapter, title, subheading
        if "/" in text or "\\" in text:
            first, after = re.split(r"[\\/]+", text, maxsplit=1)
            first = first.strip()
            after = after.strip()
            first_match = re.match(r"^(\d+(?:\.\d+)*)(?:\.)?(?:\s+(.+))?$", first)
            if first_match:
                chapter = first_match.group(1)
                title = (first_match.group(2) or "").strip()
            if after:
                subheading = after
        else:
            inline_match = re.match(r"^(\d+(?:\.\d+)*)(?:\.)?(?:\s+(.+))?$", text.strip())
            if inline_match:
                chapter = inline_match.group(1)
                title = (inline_match.group(2) or "").strip()
        return chapter, title, subheading

    def _build_detail(action: str, src: str, instruction: str, item_type: str = "") -> str:
        instruction_core = (instruction or "").split("|", 1)[0].strip()
        tail_title_match = re.search(r"(?:^|\|)\s*title\s*=\s*([^|]+)", instruction or "", re.IGNORECASE)
        tail_index_match = re.search(r"(?:^|\|)\s*index\s*=\s*([^|]+)", instruction or "", re.IGNORECASE)
        tail_title = (tail_title_match.group(1) if tail_title_match else "").strip()
        tail_index = (tail_index_match.group(1) if tail_index_match else "").strip()
        src_base = os.path.basename(src) if src else ""
        if action == "Append text":
            return src
        if action == "Extract chapter":
            chapter, title, subheading = _parse_chapter_parts(instruction_core)
            parts = [f"chapter {chapter}"] if chapter else []
            if title:
                parts.append(f"title {title}")
            if subheading:
                parts.append(f"subheading {subheading}")
            suffix = f" ({', '.join(parts)})" if parts else ""
            return f"{src_base}{suffix}".strip()
        if action == "Extract figure":
            label_match = re.search(r"\b(Figure)\b.*", instruction_core, re.IGNORECASE)
            label = label_match.group(0).strip() if label_match else ""
            parts = []
            if label:
                parts.append(label)
            elif item_type == "figure":
                parts.append("Figure")
            if tail_title:
                parts.append(f"title={tail_title}")
            if tail_index:
                parts.append(f"index={tail_index}")
            if parts:
                return f"{src_base} ({', '.join(parts)})".strip()
            return src_base
        if action == "Extract table":
            label_match = re.search(r"\b(Table)\b.*", instruction_core, re.IGNORECASE)
            label = label_match.group(0).strip() if label_match else ""
            parts = []
            if label:
                parts.append(label)
            elif item_type == "table":
                parts.append("Table")
            if tail_title:
                parts.append(f"title={tail_title}")
            if tail_index:
                parts.append(f"index={tail_index}")
            if parts:
                return f"{src_base} ({', '.join(parts)})".strip()
            return src_base
        if action == "Extract all":
            return src_base
        if action == "Extract PDF images":
            return src_base
        return src_base or instruction_core

    def _parse_instruction_tail_options(raw_instruction: str) -> tuple[str, dict[str, str], str]:
        """
        Parse optional tail tokens from mapping instruction:
        <core>|title=...|index=...
        """
        raw = (raw_instruction or "").strip()
        if not raw:
            return "", {}, ""
        parts = [p.strip() for p in raw.split("|")]
        core = parts[0]
        options: dict[str, str] = {}
        for token in parts[1:]:
            if not token:
                continue
            if "=" not in token:
                return core, options, f"無效參數語法: {token}"
            key_raw, value_raw = token.split("=", 1)
            key = key_raw.strip().lower()
            value = value_raw.strip()
            if key in {"title", "index"}:
                options[key] = value
        return core, options, ""

    for row_num, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        def _cell(idx: int) -> str:
            if idx is None or idx < 0 or idx >= len(row):
                return ""
            return str(row[idx]).strip() if row[idx] is not None else ""

        src_name = _cell(col_idx.get("source", 0))
        instruction = _cell(col_idx.get("operation", 1))
        out_rel = _cell(col_idx.get("out_path", 2))
        out_rel_normalized = _normalize_output_rel_path(out_rel)
        out_name = _cell(col_idx.get("out_name", 3))
        template_name = _cell(col_idx.get("template", 4))
        insert_label = _cell(col_idx.get("insert", 5))
        item_type = _normalize_item_type(_cell(col_idx.get("item_type", -1)))
        include_title, include_title_error = _parse_include_title(_cell(col_idx.get("include_title", -1)))

        action_label = _guess_action(instruction, item_type=item_type)
        detail_label = _build_detail(action_label, src_name, instruction, item_type=item_type)
        if include_title_error:
            _log("error", include_title_error, row_num, action_label, detail_label)
            continue
        if not instruction and item_type not in {"pdf_image", "add_text"}:
            _log("error", "缺失操作", row_num, action_label, detail_label)
            continue
        if not out_name:
            _log("error", "缺少輸出文件檔名", row_num, action_label, detail_label)
            continue
        if item_type != "add_text" and instruction.lower() != "add text" and not src_name:
            _log("error", "缺少輸入文件檔名", row_num, action_label, detail_label)
            continue
        if (item_type == "add_text" or instruction.lower() == "add text") and not src_name:
            _log("error", "Add Text 需要文字內容", row_num, action_label, detail_label)
            continue
        if insert_label and not template_name:
            _log("warn", "由於模板文件為空，插入段落將被忽略。", row_num)

        template_path = None
        if template_name:
            template_path = _find_file(task_files_dir, template_name)
            if not template_path:
                _log("error", f"未找到模板文件: {template_name}", row_num, action_label, detail_label)
                continue

        if template_path:
            if template_path not in parsed_cache:
                parsed = parse_template_paragraphs(template_path)
                index_map, last_idx = _build_template_index_map(parsed)
                parsed_cache[template_path] = (parsed, index_map, last_idx)
            parsed, index_map, last_idx = parsed_cache[template_path]
        else:
            parsed, index_map, last_idx = [], {}, None

        target_idx = None
        if template_path:
            target_key = _normalize_match(insert_label)
            target_idx = index_map.get(target_key)
            if target_idx is None:
                if insert_label:
                    if last_idx is not None:
                        target_idx = last_idx
                        _log("error", f"插入段落 '{insert_label}' 未找到", row_num, action_label, detail_label)
                    else:
                        target_idx = 0
                        _log("error", f"插入段落 '{insert_label}' 未找到", row_num, action_label, detail_label)
                else:
                    target_idx = last_idx if last_idx is not None else 0

        output_dir_full = os.path.join(output_dir, out_rel_normalized) if out_rel_normalized else output_dir
        output_path = os.path.join(output_dir_full, out_name)
        if validate_only and out_rel_normalized and not os.path.isdir(output_dir_full):
            _log("warn", f"輸出資料夾不存在: {out_rel}", row_num, action_label, detail_label)
        if output_path in output_template_map and output_template_map[output_path] != template_path:
            _log("error", f"output uses different templates: {out_name}", row_num, action_label, detail_label)
            continue
        output_template_map[output_path] = template_path

        group_key = (output_path, template_path)
        if group_key not in groups:
            groups[group_key] = {"steps": [], "parsed": parsed, "template": template_path}

        if item_type == "add_text":
            if instruction and instruction.lower() != "add text":
                _log("error", f"類型 Add Text 時，操作欄僅支援留白或 Add Text: {instruction}", row_num, action_label, detail_label)
                continue
            params = {"text": src_name}
            if template_path is not None:
                params["template_index"] = target_idx
                params["template_mode"] = "insert_after"
            params["mapping_row"] = row_num
            groups[group_key]["steps"].append({"type": "insert_text", "params": params})
            _log("info", f"append text into {out_name}", row_num)
            continue

        if instruction.lower() == "add text":
            params = {"text": src_name}
            if template_path is not None:
                params["template_index"] = target_idx
                params["template_mode"] = "insert_after"
            params["mapping_row"] = row_num
            groups[group_key]["steps"].append({"type": "insert_text", "params": params})
            _log("info", f"append text into {out_name}", row_num)
            continue

        tf_kind = None
        tf_subtitle = None
        tf_label = None
        tf_chapter = ""
        tf_chapter_title = None
        forced_kind = item_type or ""
        instruction_core = instruction
        tail_options: dict[str, str] = {}
        tail_error = ""
        if "|" in instruction and "=" in instruction.split("|", 1)[1]:
            instruction_core, tail_options, tail_error = _parse_instruction_tail_options(instruction)
            if tail_error:
                _log("error", tail_error, row_num, action_label, detail_label)
                continue

        def _parse_table_head(head_text: str) -> tuple[str, str | None, str | None]:
            parsed_chapter = ""
            parsed_title = None
            parsed_subtitle = None
            if not head_text:
                return parsed_chapter, parsed_title, parsed_subtitle
            # Accept "/", "\" and "|" as section delimiters for user convenience.
            head_parts = [p.strip() for p in re.split(r"[\\/|]+", head_text) if p.strip()]
            if not head_parts:
                head_parts = [head_text.strip()]
            inline_match = re.match(r"^(\d+(?:\.\d+)*)(?:\.)?(?:\s+(.+))?$", head_parts[0])
            if inline_match:
                parsed_chapter = inline_match.group(1).rstrip(".")
                inline_title = (inline_match.group(2) or "").strip()
                if inline_title:
                    parsed_title = inline_title
                    if len(head_parts) > 1:
                        parsed_subtitle = " ".join(head_parts[1:]).strip() or None
                else:
                    if len(head_parts) > 1:
                        parsed_title = head_parts[1].strip() or None
                    if len(head_parts) > 2:
                        parsed_subtitle = " ".join(head_parts[2:]).strip() or None
            else:
                parsed_subtitle = " ".join(head_parts).strip() or None
            return parsed_chapter, parsed_title, parsed_subtitle

        label_match = re.search(r"\b(Table|Figure)\b.*", instruction_core, re.IGNORECASE)
        if label_match:
            tf_label = instruction_core[label_match.start():].strip()
            tf_kind = "table" if tf_label.lower().startswith("table") else "figure"
            head = instruction_core[:label_match.start()].strip().rstrip("|").strip().strip(",，\u3001")
            tf_chapter, tf_chapter_title, tf_subtitle = _parse_table_head(head)
            if forced_kind and forced_kind != tf_kind:
                _log(
                    "error",
                    f"類型欄位與操作內容不一致: type={forced_kind}, label={tf_kind}",
                    row_num,
                    action_label,
                    detail_label,
                )
                continue
        elif forced_kind in {"figure", "table"}:
            tf_kind = forced_kind
            tf_label = ""
            head = instruction_core.strip().strip(",，\u3001")
            parsed_chapter, parsed_title, parsed_subtitle = _parse_chapter_parts(head)
            tf_chapter = parsed_chapter
            tf_chapter_title = parsed_title or None
            tf_subtitle = parsed_subtitle or None
        elif "title" in tail_options or "index" in tail_options:
            _log(
                "error",
                "使用 title/index 參數時必須指定 Figure 或 Table 標籤",
                row_num,
                action_label,
                detail_label,
            )
            continue

        if tf_kind:
            infile, resolve_error = _resolve_input_file(task_files_dir, src_name)
            if not infile:
                _log("error", f"來源檔案解析失敗: {resolve_error}", row_num, action_label, detail_label)
                continue
            params = {
                "input_file": infile,
                "target_chapter_section": tf_chapter,
                "include_caption": include_title,
            }
            if tf_chapter_title:
                params["target_chapter_title"] = tf_chapter_title
            if tf_subtitle:
                params["target_subtitle"] = tf_subtitle
            if tf_kind == "table":
                params["target_caption_label"] = tf_label
                option_title = (tail_options.get("title") or "").strip()
                if option_title:
                    params["target_table_title"] = option_title
                option_index_raw = (tail_options.get("index") or "").strip()
                if option_index_raw:
                    try:
                        option_index = int(option_index_raw)
                    except ValueError:
                        _log("error", f"index 必須是正整數: {option_index_raw}", row_num, action_label, detail_label)
                        continue
                    if option_index <= 0:
                        _log("error", f"index 必須大於 0: {option_index}", row_num, action_label, detail_label)
                        continue
                    params["target_table_index"] = option_index
                if not tf_label and not option_title and "target_table_index" not in params:
                    _log(
                        "error",
                        "Table 擷取至少需提供 caption、title 或 index 其中之一",
                        row_num,
                        action_label,
                        detail_label,
                    )
                    continue
                step_type = "extract_specific_table_from_word"
                hint_parts = []
                if tf_label:
                    hint_parts.append(tf_label)
                if option_title:
                    hint_parts.append(f"title={option_title}")
                if "target_table_index" in params:
                    hint_parts.append(f"index={params['target_table_index']}")
                target_hint = ", ".join(hint_parts)
                _log("info", f"extract table: {src_name} ({target_hint})", row_num)
            else:
                params["target_caption_label"] = tf_label
                option_title = (tail_options.get("title") or "").strip()
                if option_title:
                    params["target_figure_title"] = option_title
                option_index_raw = (tail_options.get("index") or "").strip()
                if option_index_raw:
                    try:
                        option_index = int(option_index_raw)
                    except ValueError:
                        _log("error", f"index 必須是正整數: {option_index_raw}", row_num, action_label, detail_label)
                        continue
                    if option_index <= 0:
                        _log("error", f"index 必須大於 0: {option_index}", row_num, action_label, detail_label)
                        continue
                    params["target_figure_index"] = option_index
                if not tf_label and not option_title and "target_figure_index" not in params:
                    _log(
                        "error",
                        "Figure 擷取至少需提供 caption、title 或 index 其中之一",
                        row_num,
                        action_label,
                        detail_label,
                    )
                    continue
                step_type = "extract_specific_figure_from_word"
                hint_parts = []
                if tf_label:
                    hint_parts.append(tf_label)
                if option_title:
                    hint_parts.append(f"title={option_title}")
                if "target_figure_index" in params:
                    hint_parts.append(f"index={params['target_figure_index']}")
                target_hint = ", ".join(hint_parts)
                _log("info", f"extract figure: {src_name} ({target_hint})", row_num)
            if template_path is not None:
                params["template_index"] = target_idx
                params["template_mode"] = "insert_after"
            params["mapping_row"] = row_num
            groups[group_key]["steps"].append({"type": step_type, "params": params})
            continue

        if item_type == "pdf_image":
            instruction_text = (instruction or "").strip().lower()
            if instruction_text not in {"", "all pages", "all", "pages"}:
                _log("error", f"PDF Image 僅支援留白、All Pages 或 All: {instruction}", row_num, action_label, detail_label)
                continue
            infile, resolve_error = _resolve_input_file(task_files_dir, src_name)
            if not infile:
                _log("error", f"來源檔案解析失敗: {resolve_error}", row_num, action_label, detail_label)
                continue
            if not str(infile).lower().endswith(".pdf"):
                _log("error", f"PDF Image 類型僅支援 PDF 檔案: {src_name}", row_num, action_label, detail_label)
                continue

            params = {"input_file": infile}
            if template_path is not None:
                params["template_index"] = target_idx
                params["template_mode"] = "insert_after"
            params["mapping_row"] = row_num
            groups[group_key]["steps"].append({"type": "extract_pdf_pages_as_images", "params": params})
            _log("info", f"extract pdf pages as images: {src_name}", row_num)
            continue

        is_all = instruction.lower() == "all"
        chapter_match = re.match(r"^([0-9]+(?:\.[0-9]+)*)(?:.*)", instruction_core)
        if not is_all and not chapter_match:
            _log("error", f"unsupported operation: {instruction_core}", row_num, action_label, detail_label)
            continue

        infile, resolve_error = _resolve_input_file(task_files_dir, src_name)
        if not infile:
            _log("error", f"來源檔案解析失敗: {resolve_error}", row_num, action_label, detail_label)
            continue

        if is_all:
            params = {"input_file": infile}
            step_type = "extract_word_all_content"
            _log("info", f"extract all: {src_name}", row_num)
        else:
            chapter = chapter_match.group(1)
            params = {"input_file": infile, "target_chapter_section": chapter}
            # Align with flow defaults
            params["ignore_toc"] = True
            params["ignore_header_footer"] = True
            params["subheading_strict_match"] = True
            params["explicit_end_title"] = ""
            params["hide_chapter_title"] = not include_title

            split_pattern = r"[\\/]+"
            has_split = re.search(split_pattern, instruction_core)
            if has_split:
                first, after = re.split(split_pattern, instruction_core, maxsplit=1)
                first = first.strip()
                after = after.strip()
                first_match = re.match(r"^(\d+(?:\.\d+)*)(?:\.)?(?:\s+(.+))?$", first)
                title_inline = ""
                if first_match:
                    chapter = first_match.group(1)
                    params["target_chapter_section"] = chapter
                    title_inline = (first_match.group(2) or "").strip()
                if after:
                    if title_inline:
                        params["target_chapter_title"] = title_inline
                    params["use_chapter_title"] = True
                    params["target_subtitle"] = after
                    if title_inline:
                        _log(
                            f"Extract chapter: {src_name} (chapter {chapter}, title {title_inline}, subheading {after})"
                            , row_num
                        )
                    else:
                        _log(f"Extract chapter: {src_name} (chapter {chapter}, subheading {after})", row_num)
                else:
                    if title_inline:
                        _log(f"Extract chapter: {src_name} (chapter {chapter}, title {title_inline})", row_num)
                    else:
                        _log(f"Extract chapter: {src_name} (chapter {chapter})", row_num)
                if "use_chapter_title" not in params:
                    params["use_chapter_title"] = False
            else:
                inline_match = re.match(r"^(\d+(?:\.\d+)*)(?:\.)?(?:\s+(.+))?$", instruction_core.strip())
                title_inline = ""
                if inline_match:
                    chapter = inline_match.group(1)
                    params["target_chapter_section"] = chapter
                    title_inline = (inline_match.group(2) or "").strip()
                params["use_chapter_title"] = False
                if title_inline:
                    params["target_chapter_title"] = title_inline
                    _log(f"Extract chapter: {src_name} (chapter {chapter}, title {title_inline})", row_num)
                else:
                    _log(f"Extract chapter: {src_name} (chapter {chapter})", row_num)
            step_type = "extract_word_chapter"
        if template_path is not None:
            params["template_index"] = target_idx
            params["template_mode"] = "insert_after"
        params["mapping_row"] = row_num
        groups[group_key]["steps"].append({"type": step_type, "params": params})

    for (output_path, template_path), payload in groups.items():
        if validate_only:
            workflow_log = []
            has_error = False
            for idx, step in enumerate(payload.get("steps", []), start=1):
                params = step.get("params", {})
                row_no = params.get("mapping_row")
                err_msg = row_errors.get(row_no, [])
                status = "error" if err_msg else "ok"
                if status == "error":
                    has_error = True
                workflow_log.append(
                    {
                        "step": idx,
                        "type": step.get("type"),
                        "params": params,
                        "status": status,
                        "error": err_msg[0] if err_msg else "",
                    }
                )
            run_logs.append(
                {
                    "output": os.path.relpath(output_path, output_dir).replace("\\", "/"),
                    "template": os.path.relpath(template_path, task_files_dir).replace("\\", "/") if template_path else None,
                    "steps": payload.get("steps", []),
                    "workflow_log": workflow_log,
                    "status": "error" if has_error else "ok",
                }
            )
            continue

        if validate_extract_only:
            workdir = os.path.join(tempfile.gettempdir(), f"mapping_validate_{uuid.uuid4().hex[:8]}")
            os.makedirs(workdir, exist_ok=True)
            template_cfg = None
            if template_path:
                template_cfg = {
                    "path": template_path,
                    "paragraphs": payload.get("parsed") or [],
                    "default_mode": "insert_after",
                }
            try:
                workflow_result = run_workflow(payload.get("steps", []), workdir=workdir, template=template_cfg)
                workflow_log = workflow_result.get("log_json", [])
                has_error = any((entry.get("status") == "error") for entry in workflow_log if isinstance(entry, dict))
                run_logs.append(
                    {
                        "output": os.path.relpath(output_path, output_dir).replace("\\", "/"),
                        "template": os.path.relpath(template_path, task_files_dir).replace("\\", "/") if template_path else None,
                        "steps": payload.get("steps", []),
                        "workflow_log": workflow_log,
                        "status": "error" if has_error else "ok",
                    }
                )
            except Exception as e:
                logs.append(f"ERROR: 驗證擷取參數失敗: {os.path.basename(output_path)} :: {e}")
                run_logs.append(
                    {
                        "output": os.path.relpath(output_path, output_dir).replace("\\", "/"),
                        "template": os.path.relpath(template_path, task_files_dir).replace("\\", "/") if template_path else None,
                        "steps": payload.get("steps", []),
                        "status": "error",
                        "error": str(e),
                        "workflow_log": [],
                    }
                )
            finally:
                shutil.rmtree(workdir, ignore_errors=True)
            continue

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        workdir = os.path.join(tempfile.gettempdir(), f"mapping_{uuid.uuid4().hex[:8]}")
        os.makedirs(workdir, exist_ok=True)
        template_cfg = None
        if template_path:
            template_cfg = {
                "path": template_path,
                "paragraphs": payload.get("parsed") or [],
                "default_mode": "insert_after",
            }
        try:
            workflow_result = run_workflow(payload.get("steps", []), workdir=workdir, template=template_cfg)
            for entry in workflow_result.get("log_json", []):
                if entry.get("status") == "error":
                    step_type = entry.get("type") or "step"
                    logs.append(
                        f"WF_ERROR: {os.path.basename(output_path)} {step_type}: {entry.get('error') or 'unknown error'}"
                    )
            result_path = workflow_result.get("result_docx") or os.path.join(workdir, "result.docx")
            titles_to_hide = collect_titles_to_hide(workflow_result.get("log_json", []))
            if DEFAULT_APPLY_FORMATTING and DEFAULT_DOCUMENT_FORMAT_KEY != "none":
                preset = DOCUMENT_FORMAT_PRESETS.get(DEFAULT_DOCUMENT_FORMAT_KEY) or DOCUMENT_FORMAT_PRESETS.get("default", {})
                apply_basic_style(
                    result_path,
                    western_font=preset.get("western_font") or "",
                    east_asian_font=preset.get("east_asian_font") or "",
                    font_size=int(preset.get("font_size") or 12),
                    line_spacing=DEFAULT_LINE_SPACING,
                    space_before=int(preset.get("space_before") or 6),
                    space_after=int(preset.get("space_after") or 6),
                )
            if not SKIP_DOCX_CLEANUP:
                remove_hidden_runs(result_path, preserve_texts=titles_to_hide)
                hide_paragraphs_with_text(result_path, titles_to_hide)
            shutil.copyfile(result_path, output_path)
            outputs.append(output_path)
            run_logs.append(
                {
                    "output": os.path.relpath(output_path, output_dir).replace("\\", "/"),
                    "template": os.path.relpath(template_path, task_files_dir).replace("\\", "/") if template_path else None,
                    "steps": payload.get("steps", []),
                    "workflow_log": workflow_result.get("log_json", []),
                    "status": "ok",
                }
            )
        except Exception as e:
            logs.append(f"Output failed: {os.path.basename(output_path)} ({e})")
            run_logs.append(
                {
                    "output": os.path.relpath(output_path, output_dir).replace("\\", "/"),
                    "template": os.path.relpath(template_path, task_files_dir).replace("\\", "/") if template_path else None,
                    "steps": payload.get("steps", []),
                    "status": "error",
                    "error": str(e),
                }
            )

    zip_file = None
    if not validate_only and not validate_extract_only and outputs:
        zip_filename = f"mapping_outputs_{uuid.uuid4().hex[:8]}.zip"
        zip_path = os.path.join(output_dir, zip_filename)
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for output_path in outputs:
                arcname = os.path.relpath(output_path, output_dir).replace("\\", "/")
                zf.write(output_path, arcname)
        zip_file = zip_filename

    log_file = None
    if run_logs or logs:
        target_log_dir = log_dir or output_dir
        os.makedirs(target_log_dir, exist_ok=True)
        log_filename = f"mapping_log_{uuid.uuid4().hex[:8]}.json"
        log_path = os.path.join(target_log_dir, log_filename)
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump({"messages": logs, "runs": run_logs}, f, ensure_ascii=False, indent=2)
        log_file = log_filename

    return {"logs": logs, "outputs": outputs, "log_file": log_file, "zip_file": zip_file}
