import json
import shutil
from io import BytesIO
from pathlib import Path

import pytest
from flask import url_for
from openpyxl import load_workbook

from app import create_app
from app.extensions import ldap_manager


@pytest.fixture
def app(monkeypatch):
    monkeypatch.setattr(ldap_manager, "init_app", lambda app: None)
    app = create_app("testing")
    ctx = app.app_context()
    ctx.push()
    try:
        yield app
    finally:
        ctx.pop()


@pytest.fixture
def client(app):
    return app.test_client()


def test_export_flow_mapping_excel_matches_mapping_headers_and_rows(app, client) -> None:
    task_id = "flow-export-mapping-xlsx"
    task_dir = Path(app.config["TASK_FOLDER"]) / task_id
    if task_dir.exists():
        shutil.rmtree(task_dir)
    (task_dir / "files").mkdir(parents=True, exist_ok=True)
    (task_dir / "flows").mkdir(parents=True, exist_ok=True)

    flow_payload = {
        "steps": [
            {"type": "extract_word_all_content", "params": {"input_file": "source/section1.docx"}},
            {
                "type": "copy_files",
                "params": {
                    "source_dir": "assets/logo.png",
                    "dest_dir": "attachments",
                    "keywords": "logo",
                    "target_name": "logo-final",
                    "recursive_search": "true",
                },
            },
            {"type": "insert_text", "params": {"text": "Appendix Note"}},
        ],
        "template_file": "templates/Device Description_Template.docx",
        "output_filename": "test/Product Description",
    }
    (task_dir / "flows" / "FlowA.json").write_text(json.dumps(flow_payload, ensure_ascii=False), encoding="utf-8")

    with app.test_request_context():
        url = url_for("flow_crud_bp.export_flow_mapping", task_id=task_id, flow_name="FlowA")

    response = client.get(url)
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
    assert "FlowA_mapping.xlsx" in response.headers.get("Content-Disposition", "")

    wb = load_workbook(filename=BytesIO(response.data))
    ws = wb.active
    assert ws.title == "Mapping定義"

    assert [ws.cell(1, i).value for i in range(1, 9)] == [
        "輸入檔案名稱/資料夾名稱/文字內容",
        "擷取類型",
        "擷取段落",
        "包含標題",
        "輸出路徑",
        "輸出檔案名稱",
        "模板文件",
        "插入段落名稱",
    ]
    assert ws.cell(1, 1).fill.fill_type == "solid"
    assert ws.cell(1, 1).font.bold is True
    assert [ws.cell(2, i).value for i in range(1, 9)] == [
        "source\\section1.docx",
        "All",
        "All",
        "Y",
        "test",
        "Product Description.docx",
        "Device Description_Template.docx",
        None,
    ]
    assert [ws.cell(3, i).value for i in range(1, 9)] == [
        "assets\\logo.png",
        "Copy File",
        "logo",
        None,
        "attachments",
        "logo-final",
        None,
        None,
    ]
    assert [ws.cell(4, i).value for i in range(1, 9)] == [
        "Appendix Note",
        "Add Text",
        "Add Text",
        None,
        "test",
        "Product Description.docx",
        "Device Description_Template.docx",
        None,
    ]
    assert ws.cell(2, 1).alignment.horizontal == "left"
    assert ws.cell(2, 1).fill.fgColor.rgb == ws.cell(1, 1).fill.fgColor.rgb
    assert ws.cell(2, 6).fill.fgColor.rgb == ws.cell(1, 6).fill.fgColor.rgb


def test_export_flow_mapping_excel_chapter_operation_includes_title_and_subtitle(app, client) -> None:
    task_id = "flow-export-mapping-chapter-title"
    task_dir = Path(app.config["TASK_FOLDER"]) / task_id
    if task_dir.exists():
        shutil.rmtree(task_dir)
    (task_dir / "files").mkdir(parents=True, exist_ok=True)
    (task_dir / "flows").mkdir(parents=True, exist_ok=True)

    flow_payload = {
        "steps": [
            {
                "type": "extract_word_chapter",
                "params": {
                    "input_file": "TD-III-011-USTAR II Knee System\\Section 1_Device Description\\Section 1_Device Description_v1.docx",
                    "target_chapter_section": "1.1.1",
                    "target_chapter_title": "General description",
                    "target_subtitle": "General description",
                },
            }
        ],
        "output_filename": "test/Product Description.docx",
        "template_file": "Device Description_Template.docx",
    }
    (task_dir / "flows" / "FlowChapter.json").write_text(json.dumps(flow_payload, ensure_ascii=False), encoding="utf-8")

    with app.test_request_context():
        url = url_for("flow_crud_bp.export_flow_mapping", task_id=task_id, flow_name="FlowChapter")

    response = client.get(url)
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.data))
    ws = wb.active
    assert ws.cell(2, 3).value == "1.1.1 General description\\General description"


def test_export_flow_mapping_excel_populates_insert_label_from_template_index(app, client, monkeypatch) -> None:
    task_id = "flow-export-mapping-insert-label"
    task_dir = Path(app.config["TASK_FOLDER"]) / task_id
    if task_dir.exists():
        shutil.rmtree(task_dir)
    files_dir = task_dir / "files"
    (files_dir / "templates").mkdir(parents=True, exist_ok=True)
    (task_dir / "flows").mkdir(parents=True, exist_ok=True)
    (files_dir / "templates" / "Template.docx").write_bytes(b"dummy")

    flow_payload = {
        "steps": [
            {
                "type": "extract_word_all_content",
                "params": {"input_file": "source.docx", "template_index": "12"},
            }
        ],
        "template_file": "templates/Template.docx",
        "output_filename": "out/result.docx",
    }
    (task_dir / "flows" / "FlowInsert.json").write_text(json.dumps(flow_payload, ensure_ascii=False), encoding="utf-8")

    monkeypatch.setattr(
        "app.blueprints.flows.flow_crud_routes.parse_template_paragraphs",
        lambda _path: [{"index": 12, "display": "a)", "text": "Product Description"}],
    )

    with app.test_request_context():
        url = url_for("flow_crud_bp.export_flow_mapping", task_id=task_id, flow_name="FlowInsert")

    response = client.get(url)
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.data))
    ws = wb.active
    assert ws.cell(2, 8).value == "a) Product Description"
